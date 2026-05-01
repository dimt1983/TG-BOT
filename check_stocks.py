"""Парсит остатки.xlsx (1С Ведомость) и сверяет с TMA-каталогом."""
import openpyxl, json, re
from pathlib import Path

STOCKS = Path("/root/projects/ai-agents-rb/BOT_TG/остатки.xlsx")
PRODUCTS_JSON = Path("/root/projects/ai-agents-rb/BOT_TG/tma_static/products.json")

def num(v):
    if v is None or v == "": return 0.0
    try: return float(v)
    except: return 0.0


def parse_stocks(fp):
    """Возвращает список листовых товаров с остатками: [{name, group, stock}, ...]"""
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw_rows = []
    for r_idx in range(10, ws.max_row + 1):
        cell = ws.cell(r_idx, 1)
        name = cell.value
        if not name: continue
        name = str(name).strip()
        row_obj = ws.row_dimensions.get(r_idx)
        outline = row_obj.outline_level if row_obj else 0
        # колонка 5 = конечный остаток
        stock = num(ws.cell(r_idx, 5).value)
        raw_rows.append({"r": r_idx, "name": name, "outline": outline, "stock": stock})

    # Листовые узлы (как в parse_stock.py из Аналитики)
    leaves = []
    MAX = 8
    group = [""] * MAX
    for i, row in enumerate(raw_rows):
        next_outline = raw_rows[i + 1]["outline"] if i + 1 < len(raw_rows) else -1
        is_leaf = next_outline <= row["outline"]
        if not is_leaf:
            group[row["outline"]] = row["name"]
            for j in range(row["outline"] + 1, MAX):
                group[j] = ""
        else:
            leaves.append({
                "name": row["name"],
                "group": " / ".join(g for g in group[:row["outline"]] if g),
                "stock": row["stock"],
            })
    return leaves


def normalize_name(s):
    s = s.upper()
    s = re.sub(r"[ЁЕ]", "Е", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def main():
    leaves = parse_stocks(STOCKS)
    print(f"Товаров на складе (листовых): {len(leaves)}")
    print(f"С остатком > 0: {sum(1 for l in leaves if l['stock'] > 0)}")
    print(f"Σ единиц: {sum(l['stock'] for l in leaves):.0f}")

    # Группы
    groups = {}
    for l in leaves:
        g = l["group"].split(" / ")[0] if l["group"] else "?"
        groups[g] = groups.get(g, 0) + 1
    print(f"\nТоп-группы:")
    for g, n in sorted(groups.items(), key=lambda x: -x[1])[:10]:
        print(f"  {g}: {n}")

    # Сравнить с TMA
    data = json.loads(PRODUCTS_JSON.read_text(encoding="utf-8"))
    tma = data["products"]
    tma_names = {normalize_name(p["name"]): p for p in tma}
    stock_names = {normalize_name(l["name"]): l for l in leaves}

    # Fuzzy match — по ключевым словам (без шумовых: г, кг, артикулы)
    NOISE = {"КГ","Г","ШТ","МЛ","Л","УПАК","ПАК","ЗЕРНАХ","ЗЕРН","КОФЕ","ЧАЙ","СИРОП",
             "ARABIKA","АРАБИКА","ITALY","ИТАЛИЯ","RB","RBR","ROASTBERRY",
             "СВЕЖЕОБЖАРЕННЫЙ","УПАКОВКА","КЛАПАН","ФОЛЬГ","ЖЕЛ","ЗЕРН"}
    def kw(s):
        s = s.upper()
        s = re.sub(r"[ЁЕ]","Е",s)
        s = re.sub(r"[^\w\s]"," ",s)
        out = set()
        for w in s.split():
            if len(w) < 3: continue
            if w.isdigit(): continue
            out.add(w)
        return out - NOISE
    def score(a,b):
        ka,kb = kw(a), kw(b)
        if not ka or not kb: return 0.0
        return len(ka&kb) / max(len(ka), len(kb))

    # Greedy: для каждого TMA-товара — лучший stock-товар
    matched_pairs = []
    used_stock = set()
    for n_tma, p in tma_names.items():
        best, best_s = None, 0.0
        for n_st, l in stock_names.items():
            if n_st in used_stock: continue
            s = score(p["name"], l["name"])
            if s > best_s:
                best_s, best = s, l
        if best and best_s >= 0.45:
            matched_pairs.append((p, best, best_s))
            used_stock.add(normalize_name(best["name"]))

    # Применяем остатки
    out_of_stock = 0
    for p, l, sc in matched_pairs:
        p["stock"] = int(l["stock"])
        if l["stock"] <= 0:
            out_of_stock += 1

    print(f"\n=== Сверка (fuzzy) ===")
    print(f"Сматчено: {len(matched_pairs)} / TMA: {len(tma)}, склад: {len(leaves)}")
    print(f"С остатком 0: {out_of_stock}")
    only_tma = [p for p in tma if not any(p is mp[0] for mp in matched_pairs)]
    print(f"TMA без склада: {len(only_tma)}")

    # Сохраним отчёт
    PRODUCTS_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    report = {
        "matched": len(matched_pairs),
        "tma_total": len(tma),
        "stock_total": len(leaves),
        "only_in_tma_sample": [p["name"] for p in only_tma[:30]],
        "only_in_stock_sample": [l["name"] for l in leaves if normalize_name(l["name"]) not in used_stock][:30],
    }
    Path("/root/projects/ai-agents-rb/BOT_TG/stocks_diff.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nОтчёт расхождений: stocks_diff.json")
    print(f"\nПримеры 'только в TMA' (5):")
    for p in only_tma[:5]:
        print(f"  {p['name']}")
    print(f"\nПримеры 'только на складе' (5):")
    for l in leaves:
        if normalize_name(l["name"]) not in used_stock:
            print(f"  {l['name'][:80]} (остаток {l['stock']:.0f})")


if __name__ == "__main__":
    main()
