#!/usr/bin/env python3
"""Stock updater: fresh 1C rb-bot.xlsx -> TMA products.json.

Default flow:
  1. Fetch fresh yadisk:Roastberry/1C/rb-bot.xlsx into _diff/rb-bot.latest.xlsx.
  2. Parse the 1C report by header names, not hardcoded rows/columns.
  3. Match products by exact normalized name, then by meaningful tokens.
  4. Update stock; if a matched card was quarantined only because it was absent
     from the stock file before, bring it back.
  5. Print visible zero-stock warnings separately from hidden/quarantined zeros.

Dry-run by default:
  python3 scripts/update_stock_from_xlsx.py

Apply:
  python3 scripts/update_stock_from_xlsx.py --apply
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
PJSON = ROOT / "tma_static" / "products.json"
DIFF_DIR = ROOT / "_diff"
FETCHED_XLSX = DIFF_DIR / "rb-bot.latest.xlsx"
REMOTE = "yadisk:Roastberry/1C/rb-bot.xlsx"

FALLBACK_XLSX = [
    Path("/root/projects/ai-agents-rb/bishoprb-agent/workdir/rb-bot.xlsx"),
    ROOT / "RB-BOT.xlsx",
]

TOP_GROUPS = {
    "КОФЕ": "coffee",
    "DRIP": "drip",
    "САХАР": "other",
    "КОНЦЕНТРАТЫ": "syrup",
    "СИРОПЫ": "syrup",
    "МОЛОКО": "milk",
    "ЧАЙ": "tea",
}

SECTION_GROUPS = {
    "ESPRESSO RBR": "coffee",
    "FILTER RBR": "coffee",
    "ROASTBERRY": "coffee",
    "1000 \u0413": "coffee",
    "200 \u0413": "coffee",
    "200 \u0413 BLACK": "coffee",
    "200 \u0413 BOR\u0429": "coffee",
    "\u041f\u041e\u0421\u0423\u0414\u0410 \u0418 \u0410\u041a\u0421\u0415\u0421\u0421\u0423\u0410\u0420\u042b": "other",
    "BARLINE": "syrup",
    "BOTANIKA": "syrup",
    "VEDRENNE": "syrup",
    "\u041a\u041e\u0420\u0414\u0418\u0410\u041b": "syrup",
    "\u041c\u041e\u041d\u0418\u041d": "syrup",
    "ALTHAUS": "tea",
    "NIKTEA": "tea",
    "RBR TEA": "tea",
    "RESTORANICA": "tea",
    "ICEDTEA-400": "tea",
    "RESTA - 650": "tea",
    "TOGO - 350": "tea",
}

CATEGORY_GROUPS = {
    "tea": {"tea"},
    "syrup": {"syrup"},
    "milk": {"milk"},
    "other": {"other", "syrup", "milk"},
    "drip": {"drip", "coffee"},
    "espresso": {"coffee"},
    "filter": {"coffee"},
    "retail": {"coffee"},
    "black": {"coffee"},
    "borshch": {"coffee"},
    "cascara": {"coffee", "tea"},
}

SKIP_CATEGORIES = {"consulting", "sets"}
FIXED_STOCK_CATEGORIES = {"consulting": 1}

UNIT_RE = re.compile(
    r"(?:\b\d+(?:[.,]\d+)?\s*(?:\u043a\u0433|\u0433|\u0433\u0440|\u043c\u043b|\u043b|ml|kg|gr|\u0448\u0442|\u0448\u0442\u0443\u043a|\u043f\u0430\u043a\.?|\u043f\u0430\u043a\u0435\u0442\w*)\b)"
    r"|(?:\d+\s*[x\u0445]\s*\d+(?:[.,]\d+)?)"
    r"|(?:\(\s*\d+\s*\))|(?:\b\d+\s*\u043f\u0430\u043a\b)"
    r"|(?:\u0444\u0438\u043b\u044c\u0442\u0440[- ]?\u043f\u0430\u043a\u0435\u0442\w*)|(?:\u0432\u0430\u043b\u043e\u043c)",
    re.IGNORECASE,
)


KNOWN_GROUPS = {
    "1000 г", "200 г", "200 г black", "200 г borщ", "200 г борщ",
    "8 шт упаоква", "8 шт упаковка", "по штучно rbr (свои полностью)",
    "icedtea-400", "resta-650", "togo - 350", "молоко животное", "молоко растительное",
}

NOISE = {
    "ЧАЙ", "ЧАИ", "ЧЕРНЫЙ", "ЧЕРН", "ЧЕРН.", "ЗЕЛЕНЫЙ", "ЗЕЛЕН", "ЗЕЛ",
    "БЕЛЫЙ", "КРАСНЫЙ", "ФРУКТОВЫЙ", "ФРУКТ", "ТРАВЯНОЙ", "ТРАВ",
    "ЛИСТОВОЙ", "ЛИСТ", "БАЙХОВЫЙ", "БАЙХ", "ПАКЕТ", "ПАК", "ПАКЕТИР",
    "ПАКЕТИРОВАННЫЙ", "ПИРАМИДКИ", "ПИРАМИДКА", "АРОМАТ", "АРОМАТИЗ",
    "НАПИТОК", "ДЛЯ", "ЧАЙНИКА", "Д", "NEW", "НОВ", "СИРОП", "ТОППИНГ", "СИРОПЫ",
    "КОРДИАЛ", "ПАРФЮМ", "БУТ", "СТЕКЛ", "БАНКА", "БАН", "МОЛОКО",
    "КОФЕ", "МОЛОТЫЙ", "МОЛОТ", "ЭСПРЕССО", "ФИЛЬТР", "FILTER", "ESPRESSO",
    "RBR", "ROASTBERRY", "ALTHAUS", "АЛЬТХАУС", "АЛЬТАУС", "NIKTEA",
    "TASTEABREW", "BARLINE", "BOTANIKA", "HERBARISTA", "SWEETSHOT",
    "BARBACKS", "CLAVIS", "МЛ", "Л", "КГ", "Г", "ШТ", "Х", "X",
    "НА", "ПО", "С", "И", "В", "ИЗ", "ОТ", "СО", "ОСНОВЕ",
}

ARTICLE_RE = re.compile(r"^(?:ЧА|ЧН|NP|NR|АР|КФ|МК|МЛ|ТП|СР)?\d{1,5}[A-ZА-Я-]*$")


SIZE_RE = re.compile(
    r"(?<!\d)(\d+(?:[.,]\d+)?)\s*"
    r"(\u043a\u0433|\u0433\u0440|\u0433|kg|gr|g|\u043c\u043b|\u043b|ml|l)\b",
    re.IGNORECASE,
)

BRAND_ALIASES = {
    "ALTHAUS": ("ALTHAUS",),
    "BARBACKS": ("BARBACKS",),
    "BARLINE": ("BARLINE",),
    "BOTANIKA": ("BOTANIKA",),
    "CLAVIS": ("CLAVIS",),
    "HERBARISTA": ("HERBARISTA",),
    "MILLER": ("MILLER",),
    "MONIN": ("MONIN",),
    "NIKTEA": ("NIKTEA",),
    "SWEETSHOT": ("SWEETSHOT",),
    "TASTEABREW": ("TASTEABREW",),
    "VEDRENNE": ("VEDRENNE",),
}


def nspace(s: str) -> str:
    return str(s or "").replace("\xa0", " ")


def norm(s: str) -> str:
    s = nspace(s).upper().replace("Ё", "Е")
    s = s.replace("×", "Х").replace("X", "Х")
    s = re.sub(r"[«»\"'`]+", " ", s)
    s = re.sub(r"[^A-ZА-Я0-9Х.,%\s-]+", " ", s)
    s = re.sub(r"\bГР\.?\b", "Г", s)
    s = re.sub(r"\bЛ\.\b", "Л", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tokens(s: str) -> set[str]:
    out: set[str] = set()
    for raw in norm(s).replace(",", " ").replace(".", " ").split():
        for w in re.split(r"[-/]", raw.strip("-")):
            if len(w) < 3 or w.isdigit() or ARTICLE_RE.match(w) or w in NOISE:
                continue
            out.add(w)
    return out


def brands(s: str) -> set[str]:
    ns = norm(s)
    found: set[str] = set()
    for brand, aliases in BRAND_ALIASES.items():
        for alias in aliases:
            alias_norm = norm(alias)
            if re.search(rf"(?<!\w){re.escape(alias_norm)}(?!\w)", ns):
                found.add(brand)
                break
    return found


def size_keys(s: str) -> set[str]:
    found: set[str] = set()
    for m in SIZE_RE.finditer(nspace(s).lower()):
        value = float(m.group(1).replace(",", "."))
        unit = m.group(2)
        if unit in {"kg", "\u043a\u0433"}:
            found.add(f"g:{value * 1000:g}")
        elif unit in {"g", "gr", "\u0433", "\u0433\u0440"}:
            found.add(f"g:{value:g}")
        elif unit in {"l", "\u043b"}:
            found.add(f"ml:{value * 1000:g}")
        elif unit in {"ml", "\u043c\u043b"}:
            found.add(f"ml:{value:g}")
    return found


def is_leaf(name: str) -> bool:
    low = norm(name).lower()
    if low in KNOWN_GROUPS:
        return False
    if not UNIT_RE.search(name):
        return False
    meaningful = [w for w in re.split(r"\s+", name) if re.search(r"[A-Za-zА-Яа-я]{3,}", w)]
    return len(meaningful) >= 1


def json_dumps(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, indent=2)


def readable_file(path: Path) -> bool:
    try:
        return path.exists() and path.is_file() and path.stat().st_size > 0
    except OSError:
        return False


def fetch_xlsx() -> Path:
    DIFF_DIR.mkdir(parents=True, exist_ok=True)
    tmp = DIFF_DIR / "rb-bot.latest.tmp.xlsx"
    try:
        if tmp.exists():
            tmp.unlink()
        r = subprocess.run(
            ["rclone", "copyto", REMOTE, str(tmp)],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if r.returncode == 0 and tmp.exists() and tmp.stat().st_size > 0:
            tmp.replace(FETCHED_XLSX)
            return FETCHED_XLSX
        print(f"[warn] rclone failed, fallback to local copy: {r.stderr.strip()[:300]}", file=sys.stderr)
    except Exception as e:
        print(f"[warn] rclone unavailable, fallback to local copy: {e}", file=sys.stderr)

    candidates = [p for p in [FETCHED_XLSX, *FALLBACK_XLSX] if readable_file(p)]
    if not candidates:
        raise FileNotFoundError("No rb-bot.xlsx source found")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_header(ws) -> tuple[int, int, int]:
    name_col = qty_col = None
    header_row = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [norm(v) for v in row]
        if "НОМЕНКЛАТУРА" not in vals:
            continue
        header_row = row_idx
        name_col = vals.index("НОМЕНКЛАТУРА")
        for next_row in ws.iter_rows(min_row=row_idx, max_row=min(ws.max_row, row_idx + 2), values_only=True):
            nvals = [norm(v) for v in next_row]
            if "КОНЕЧНЫЙ ОСТАТОК" in nvals:
                qty_col = nvals.index("КОНЕЧНЫЙ ОСТАТОК")
                break
        if qty_col is None and "КОЛИЧЕСТВО" in vals:
            qty_col = vals.index("КОЛИЧЕСТВО")
        if qty_col is not None:
            return header_row, name_col, qty_col
    raise RuntimeError("Cannot find Номенклатура / Конечный остаток headers")


def load_excel_items(xlsx: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, name_col, qty_col = find_header(ws)
    current_group = None
    items: list[dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        name = row[name_col] if name_col < len(row) else None
        if not isinstance(name, str) or not name.strip():
            continue
        raw = nspace(name).strip()
        if norm(raw) in {"КОНЕЧНЫЙ ОСТАТОК", "КОЛИЧЕСТВО"}:
            continue
        qty_raw = row[qty_col] if qty_col < len(row) else None
        try:
            qty_f = float(qty_raw) if qty_raw is not None else 0.0
        except (TypeError, ValueError):
            qty_f = 0.0
        qty: int | float = int(qty_f) if qty_f.is_integer() else qty_f

        group_hit = TOP_GROUPS.get(norm(raw))
        if group_hit:
            current_group = group_hit
            continue
        section_hit = SECTION_GROUPS.get(norm(raw))
        if section_hit:
            current_group = section_hit
            continue
        if not is_leaf(raw):
            continue
        items.append({
            "name": raw,
            "qty": qty,
            "group": current_group or "unknown",
            "row": row_idx,
            "norm": norm(raw),
            "tokens": tokens(raw),
        })
    return items


def product_candidates(p: dict[str, Any]) -> list[str]:
    name = p.get("name") or ""
    out = [name]
    for f in p.get("fasovka") or []:
        size = f.get("size") or ""
        if size:
            out.extend([f"{name} {size}", f"{name} - {size}", f"{name}, {size}"])
    return out


def allowed_items_for_product(p: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups = CATEGORY_GROUPS.get(p.get("category"), set(TOP_GROUPS.values()))
    return [it for it in items if it["group"] in groups]


def semantic_score(product_toks: set[str], item_toks: set[str]) -> float:
    if not product_toks or not item_toks:
        return 0.0
    overlap = len(product_toks & item_toks)
    precision = overlap / len(product_toks)
    recall = overlap / len(item_toks)
    return precision * 0.72 + recall * 0.28


def match_product(p: dict[str, Any], items: list[dict[str, Any]], by_norm: dict[str, dict[str, Any]]) -> tuple[dict[str, Any] | None, str, float, list[tuple[float, str]]]:
    allowed = allowed_items_for_product(p, items)
    product_text = " ".join(product_candidates(p))
    p_brands = brands(product_text)
    p_sizes = size_keys(product_text)
    for cand in product_candidates(p):
        hit = by_norm.get(norm(cand))
        if hit and hit in allowed:
            return hit, "exact", 1.0, []

    p_toks = set()
    for cand in product_candidates(p):
        p_toks |= tokens(cand)
    if not p_toks:
        return None, "none", 0.0, []

    scored: list[tuple[float, dict[str, Any]]] = []
    for it in allowed:
        it_brands = brands(it["name"])
        if p_brands and it_brands and p_brands.isdisjoint(it_brands):
            continue
        it_sizes = size_keys(it["name"])
        if p_sizes and it_sizes and p_sizes.isdisjoint(it_sizes):
            continue
        s = semantic_score(p_toks, it["tokens"])
        overlap = len(p_toks & it["tokens"])
        if overlap >= 2 and s >= 0.72:
            scored.append((s, it))
    scored.sort(key=lambda x: x[0], reverse=True)
    top = [(round(s, 3), it["name"]) for s, it in scored[:3]]
    if not scored:
        return None, "none", 0.0, top

    best_s, best = scored[0]
    second_s = scored[1][0] if len(scored) > 1 else 0.0
    if best_s < 0.80:
        return None, "low_score", best_s, top
    if second_s and best_s - second_s < 0.08:
        return None, "ambiguous", best_s, top
    return best, "semantic", best_s, top


def zero_report(products: list[dict[str, Any]]) -> tuple[dict[str, int], dict[str, list[str]]]:
    all_zero: dict[str, int] = defaultdict(int)
    visible_zero: dict[str, list[str]] = defaultdict(list)
    for p in products:
        if (p.get("stock") or 0) != 0:
            continue
        cat = p.get("category") or "unknown"
        all_zero[cat] += 1
        if not p.get("hidden") and not p.get("quarantined") and not p.get("archived"):
            visible_zero[cat].append(p.get("name") or p.get("id") or "")
    return dict(all_zero), dict(visible_zero)


def _return_entry(
    p: dict[str, Any],
    hit: dict[str, Any] | None,
    method: str,
    score: float,
    top: list[tuple[float, str]],
    reason: str,
) -> dict[str, Any]:
    entry = {
        "id": p.get("id"),
        "name": p.get("name"),
        "category": p.get("category"),
        "old": p.get("stock") or 0,
        "hidden": bool(p.get("hidden")),
        "quarantined": bool(p.get("quarantined")),
        "archived": bool(p.get("archived")),
        "reason": reason,
        "method": method,
        "score": round(score, 3),
    }
    if hit:
        entry.update({
            "new": hit["qty"],
            "xlsx": hit["name"],
            "row": hit["row"],
        })
    if top:
        entry["top"] = top[:5]
    return entry


def build_return_candidates(
    products: list[dict[str, Any]],
    items: list[dict[str, Any]],
    by_norm: dict[str, dict[str, Any]],
    category: str | None = None,
) -> dict[str, Any]:
    report: dict[str, Any] = {
        "summary": {
            "can_restore": 0,
            "needs_review": 0,
            "matched_zero": 0,
            "no_match": 0,
        },
        "can_restore": [],
        "needs_review": [],
        "matched_zero": [],
        "no_match": [],
    }

    for p in products:
        if p.get("category") in SKIP_CATEGORIES:
            continue
        if category and p.get("category") != category:
            continue
        is_hidden = bool(p.get("hidden"))
        is_quar = bool(p.get("quarantined"))
        is_archived = bool(p.get("archived"))
        if not (is_hidden or is_quar):
            continue

        hit, method, score, top = match_product(p, items, by_norm)
        if not hit:
            report["no_match"].append(_return_entry(p, None, method, score, top, "no_match"))
            continue

        if not hit["qty"]:
            report["matched_zero"].append(_return_entry(p, hit, method, score, top, "matched_but_1c_stock_is_zero"))
            continue

        if is_archived:
            report["needs_review"].append(_return_entry(p, hit, method, score, top, "archived_needs_manual_review"))
        elif is_quar:
            report["can_restore"].append(_return_entry(p, hit, method, score, top, "quarantined_with_1c_stock"))
        else:
            report["needs_review"].append(_return_entry(p, hit, method, score, top, "hidden_not_quarantined"))

    for key in ("can_restore", "needs_review", "matched_zero", "no_match"):
        report["summary"][key] = len(report[key])
    return report


def print_return_candidates(report: dict[str, Any], limit: int = 40) -> None:
    summary = report.get("summary", {})
    print("\n=== Hidden/quarantined return candidates ===")
    print(f"can_restore: {summary.get('can_restore', 0)}")
    print(f"needs_review: {summary.get('needs_review', 0)}")
    print(f"matched_zero: {summary.get('matched_zero', 0)}")
    print(f"no_match: {summary.get('no_match', 0)}")

    def show(title: str, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        print(f"\n{title}:")
        for row in rows[:limit]:
            new = row.get("new", "?")
            xlsx = row.get("xlsx", "no xlsx match")
            print(
                f"  - {row.get('id')} | {row.get('old')} -> {new} | "
                f"{row.get('method')} {row.get('score')} | {row.get('name')} <= {xlsx}"
            )
        if len(rows) > limit:
            print(f"  ... and {len(rows) - limit} more")

    show("can_restore", report.get("can_restore", []))
    show("needs_review", report.get("needs_review", []))
    show("no_match", report.get("no_match", []))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="Write products.json")
    ap.add_argument("--xlsx", type=Path, default=None, help="Use explicit xlsx path")
    ap.add_argument("--skip-fetch", action="store_true", help="Use latest local/fallback xlsx without rclone")
    ap.add_argument("--no-unquarantine", action="store_true", help="Do not return matched quarantined cards")
    ap.add_argument("--report", type=Path, default=DIFF_DIR / "stock_sync_report.json")
    ap.add_argument("--return-candidates", action="store_true", help="Report hidden/quarantined products that have positive 1C stock")
    ap.add_argument("--return-candidates-only", action="store_true", help="Only print hidden/quarantined return candidates and exit")
    ap.add_argument("--category", default=None, help="Limit return-candidates report to one category, e.g. tea")
    args = ap.parse_args()

    if args.xlsx:
        xlsx = args.xlsx
    elif args.skip_fetch:
        candidates = [p for p in [FETCHED_XLSX, *FALLBACK_XLSX] if readable_file(p)]
        if not candidates:
            raise FileNotFoundError("No readable rb-bot.xlsx source found")
        xlsx = max(candidates, key=lambda p: p.stat().st_mtime)
    else:
        xlsx = fetch_xlsx()
    items = load_excel_items(xlsx)
    by_norm: dict[str, dict[str, Any]] = {}
    for it in items:
        by_norm[it["norm"]] = it

    data = json.loads(PJSON.read_text(encoding="utf-8"))
    products = data.get("products", [])

    return_candidates = build_return_candidates(products, items, by_norm, args.category)

    if args.return_candidates_only:
        report = {
            "xlsx": str(xlsx),
            "xlsx_items": len(items),
            "return_candidates": return_candidates,
        }
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json_dumps(report) + "\n", encoding="utf-8")
        print("\n=== Stock return candidate summary ===")
        print(f"xlsx: {xlsx}")
        print(f"xlsx items: {len(items)}")
        print(f"report: {args.report}")
        print_return_candidates(return_candidates)
        return 0

    changes = []
    unmatched_visible_zero = []
    matched = 0
    skipped = 0

    for p in products:
        fixed_stock = FIXED_STOCK_CATEGORIES.get(p.get("category"))
        if fixed_stock is not None:
            old_stock = p.get("stock") or 0
            was_hidden = bool(p.get("hidden"))
            was_quar = bool(p.get("quarantined"))
            if old_stock != fixed_stock or was_hidden or was_quar:
                changes.append({
                    "id": p.get("id"),
                    "name": p.get("name"),
                    "old": old_stock,
                    "new": fixed_stock,
                    "category": p.get("category"),
                    "xlsx": "fixed category stock",
                    "method": "fixed_category_stock",
                    "score": 100,
                    "restored": was_quar,
                })
                if args.apply:
                    p["stock"] = fixed_stock
                    p["hidden"] = False
                    p["quarantined"] = False
                    p["quarantined_at"] = None
                    p["quarantine_reason"] = None
            skipped += 1
            continue
        if p.get("category") in SKIP_CATEGORIES:
            continue
        hit, method, score, top = match_product(p, items, by_norm)
        if not hit:
            skipped += 1
            if (p.get("stock") or 0) == 0 and not p.get("hidden") and not p.get("quarantined") and not p.get("archived"):
                unmatched_visible_zero.append({
                    "id": p.get("id"),
                    "name": p.get("name"),
                    "category": p.get("category"),
                    "top": top,
                })
            continue

        matched += 1
        old_stock = p.get("stock") or 0
        new_stock = hit["qty"]
        was_hidden = bool(p.get("hidden"))
        was_quar = bool(p.get("quarantined"))
        should_restore = not args.no_unquarantine and new_stock and was_quar

        if old_stock != new_stock or should_restore:
            changes.append({
                "id": p.get("id"),
                "name": p.get("name"),
                "old": old_stock,
                "new": new_stock,
                "category": p.get("category"),
                "method": method,
                "score": round(score, 3),
                "xlsx": hit["name"],
                "row": hit["row"],
                "restore": bool(should_restore),
            })
            if args.apply:
                p["stock"] = new_stock
                if should_restore:
                    p["hidden"] = False
                    p["quarantined"] = False
                    p["quarantined_at"] = None
                    p["quarantine_reason"] = None

    all_zero, visible_zero = zero_report(products)
    report = {
        "xlsx": str(xlsx),
        "xlsx_items": len(items),
        "matched_products": matched,
        "skipped_products": skipped,
        "changes": changes,
        "unmatched_visible_zero": unmatched_visible_zero,
        "return_candidates": return_candidates,
        "zero_all_by_category": all_zero,
        "zero_visible_by_category": {k: len(v) for k, v in visible_zero.items()},
        "zero_visible_examples": {k: v[:15] for k, v in visible_zero.items()},
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json_dumps(report) + "\n", encoding="utf-8")

    prefix = "[APPLY]" if args.apply else "[DRY]"
    for ch in changes[:120]:
        restore = " restore" if ch["restore"] else ""
        print(f"{prefix} {ch['old']} -> {ch['new']}{restore} | {ch['category']} | {ch['name']} | <= {ch['xlsx']}")
    if len(changes) > 120:
        print(f"... and {len(changes) - 120} more changes")

    print("\n=== Stock sync summary ===")
    print(f"xlsx: {xlsx}")
    print(f"xlsx items: {len(items)}")
    print(f"matched products: {matched}")
    print(f"{'updated/restored' if args.apply else 'would update/restore'}: {len(changes)}")
    print(f"report: {args.report}")
    if args.return_candidates:
        print_return_candidates(return_candidates)

    visible_total = sum(len(v) for v in visible_zero.values())
    print("\n=== Visible zero stock check ===")
    print(f"visible zero total: {visible_total}")
    if not visible_total:
        print("  OK: no visible zero-stock products")
    for cat, names in sorted(visible_zero.items(), key=lambda kv: -len(kv[1])):
        print(f"  visible {cat}: {len(names)} | {', '.join(names[:5])}")

    print("\n=== Hidden/quarantined zero stock (informational) ===")
    print(f"all zero total (informational): {sum(all_zero.values())}")
    for cat, n in sorted(all_zero.items(), key=lambda kv: -kv[1]):
        print(f"  all {cat}: {n}")

    if args.apply and changes:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = PJSON.with_name(f"products.json.bak.stock_sync.{ts}")
        shutil.copy2(PJSON, backup)
        PJSON.write_text(json_dumps(data) + "\n", encoding="utf-8")
        print(f"\nSaved products.json; backup: {backup}")
    elif not args.apply:
        print("\nDry run only. Use --apply to write.")

    if visible_total:
        print("\nWARNING: visible zero-stock products remain. See report for details.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
