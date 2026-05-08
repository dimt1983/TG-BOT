"""Синк цен (Базовый прайс) и остатков из 'Прайс и остатки.xlsx' в TMA-каталог.

Файл — выгрузка 1С с иерархией:
  outline=0: категория (МОЛОКО / СИРОПЫ / ЧАЙ)
  outline=1: бренд (BARLINE / BOTANIKA / ALTHAUS / ...)
  outline=2: товар
Колонки:
  E (5)  — Номенклатура
  O (15) — Цена (базовый прайс, RUB)
  Q (17) — Остаток
"""
import openpyxl, json, re
from pathlib import Path

XLSX = Path("/root/projects/ai-agents-rb/BOT_TG/Прайс и остатки.xlsx")
STOCKS_XLSX = Path("/root/projects/ai-agents-rb/BOT_TG/Ведомость остатков.xlsx")
PRODUCTS = Path("/root/projects/ai-agents-rb/BOT_TG/tma_static/products.json")


def num(v):
    if v is None or v == "": return None
    try: return float(v)
    except: return None


def normalize(s):
    s = s.upper()
    s = re.sub(r"[ЁЕ]","Е",s)
    s = re.sub(r"[^\w\s]"," ",s)
    return s


_ARTICUL_RE = re.compile(r"^(ЧА|ЧАD|АР|КФ|МК|МЛ|ТП|СР)?\d{1,5}[A-ZА-Я]?$")


def kw(s):
    s = normalize(s)
    NOISE = {
        # Категории и общие слова
        "СИРОП","ТОППИНГ","СИРОПЫ","ЧАЙ","НАПИТОК","КОРДИАЛ","ПАРФЮМ",
        "ЧАИ","ЧАЯ","ЧАЙНИКОВ","ЧАЙНИКА","ЧАЙНИК","ЧАШКА","ЧАШЕК","ЧАШКУ",
        # Цвета и типы
        "ЧЕРНЫЙ","ЧЕРН","ЗЕЛЕНЫЙ","ЗЕЛ","ЗЕЛЕН","БЕЛЫЙ","БЕЛ","КРАСНЫЙ","ФРУКТ",
        "ФРУКТОВЫЙ","ФРУКТОВ","ТРАВЯНОЙ","ТРАВ","ЯГОДНЫЙ","ЯГОД",
        # Форматы
        "ЛИСТОВОЙ","ЛИСТ","БАЙХОВЫЙ","БАЙХ","ПАКЕТ","ПАК","ПАКЕТИРОВАННЫЙ",
        "ПАКЕТИР","ПАКЕТИРОВ","ПИРАМИДКА","ПИРАМИДКИ","ПИРАМ","ПИР",
        "АРОМАТ","АРОМАТИЗ","АРОМАТИЗИРОВАННЫЙ","АРОМ","АРАМАТ",
        # Бренды и сетки
        "BARLINE","BOTANIKA","ALTHAUS","АЛЬТХАУС","АЛЬТАУС",
        "NIKTEA","ROASTBERRY","RBR","HERBARISTA","SWEETSHOT","ICEDREAM",
        "GREEN","MILK","BARBACKS","CLAVIS","МОНИН","МИЛЛЕР","VEDRENNE",
        # Единицы и связки
        "Б","А","Л","КГ","Г","ШТ","МЛ","НА","СОЕВОЙ","ОСНОВЕ","ДЛЯ",
        "С","И","ИЛИ","КОФЕ","БУТ","СТЕКЛ","СТ","БАНКА","БАН",
        "NEW","НОВ","ХИТ","ЧАЙН",
        "ИЗ","ОТ","ПО","ВО","СО","ВКУС","ВКУСОМ","ОБРАЗЦА",
    }
    out = set()
    for w in s.split():
        if len(w) < 3: continue
        if w.isdigit(): continue
        if _ARTICUL_RE.match(w): continue  # артикулы вида ЧА016, ЧАD012, КФ123
        out.add(w)
    return out - NOISE


def score(a, b):
    ka, kb = kw(a), kw(b)
    if not ka or not kb: return 0.0
    return len(ka & kb) / max(len(ka), len(kb))


def parse_xlsx():
    """Прайс-лист 1С: имя в кол.5, цена в кол.15, остаток в кол.17,
    товар = outline≥2."""
    if not XLSX.exists():
        return []
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 5).value
        if not name: continue
        name = str(name).strip()
        row_obj = ws.row_dimensions.get(r)
        outline = row_obj.outline_level if row_obj else 0
        if outline < 2: continue  # только товары
        price = num(ws.cell(r, 15).value)
        stock = num(ws.cell(r, 17).value)
        if price is None: continue
        items.append({"name": name, "price": price, "stock": int(stock or 0)})
    return items


def parse_stocks_xlsx():
    """Ведомость остатков 1С: имя в кол.1, остаток в кол.5, товар = outline≥3.
    Возвращает [{name, stock}] (без цен)."""
    if not STOCKS_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(STOCKS_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name: continue
        name = str(name).strip()
        row_obj = ws.row_dimensions.get(r)
        outline = row_obj.outline_level if row_obj else 0
        if outline < 3: continue  # товары на 3-м уровне иерархии
        stock = num(ws.cell(r, 5).value)
        if stock is None: continue
        items.append({"name": name, "stock": int(stock)})
    return items


def main():
    items = parse_xlsx()
    stocks_items = parse_stocks_xlsx()
    print(f"Распарсено товаров из 'Прайс и остатки.xlsx': {len(items)}")
    print(f"Распарсено товаров из 'Ведомость остатков.xlsx': {len(stocks_items)}")

    data = json.loads(PRODUCTS.read_text(encoding="utf-8"))
    products = data["products"]

    # Из прайса берём цены и остатки, но только для не-кофе
    # (цены кофе живут в Прайсы/чистовики и подсасываются live_prices_api).
    targets = [p for p in products if p.get("category") != "coffee"]
    print(f"Целевых TMA-товаров (не-кофе) для прайса: {len(targets)}")

    # Матчим: для каждого xlsx-item ищем лучший TMA target
    pairs = []
    for src in items:
        best, best_s = None, 0.0
        for t in targets:
            s = score(src["name"], t["name"])
            if s > best_s:
                best_s, best = s, t
        if best and best_s >= 0.5:
            pairs.append((best_s, src, best))
    pairs.sort(reverse=True, key=lambda x: x[0])

    used_tma = set()
    used_xlsx = set()
    matched = 0
    price_changes = 0
    stock_changes = 0
    for s, src, t in pairs:
        if t["id"] in used_tma: continue
        if src["name"] in used_xlsx: continue
        used_tma.add(t["id"])
        used_xlsx.add(src["name"])
        # Применяем: цена идёт в первую (или единственную) фасовку
        old_price = t["fasovka"][0]["price"] if t.get("fasovka") else None
        if old_price != src["price"]:
            t["fasovka"][0]["price"] = src["price"]
            price_changes += 1
        old_stock = t.get("stock", 0)
        if old_stock != src["stock"]:
            t["stock"] = src["stock"]
            stock_changes += 1
        matched += 1

    # Второй проход — ведомость остатков. Включает кофе тоже (это и есть
    # отдельный отчёт по складам который Дмитрий шлёт раз в день).
    stock_only_changes = 0
    stock_only_matched = 0
    stock_unmatched_examples = []
    if stocks_items:
        all_targets = products  # все товары, включая кофе
        used_st_tma = set()
        used_st_xlsx = set()
        st_pairs = []
        for src in stocks_items:
            best, best_s = None, 0.0
            for t in all_targets:
                s = score(src["name"], t["name"])
                if s > best_s:
                    best_s, best = s, t
            if best and best_s >= 0.5:
                st_pairs.append((best_s, src, best))
        st_pairs.sort(reverse=True, key=lambda x: x[0])
        for s, src, t in st_pairs:
            if t["id"] in used_st_tma: continue
            if src["name"] in used_st_xlsx: continue
            used_st_tma.add(t["id"])
            used_st_xlsx.add(src["name"])
            stock_only_matched += 1
            old_stock = t.get("stock", 0)
            if old_stock != src["stock"]:
                t["stock"] = src["stock"]
                stock_only_changes += 1
        stock_unmatched_examples = [
            s["name"] for s in stocks_items if s["name"] not in used_st_xlsx
        ]

    PRODUCTS.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n=== Прайс ===")
    print(f"Сматчено: {matched} / xlsx={len(items)}, tma={len(targets)}")
    print(f"Цены обновлены: {price_changes}")
    print(f"Остатки обновлены: {stock_changes}")
    if stocks_items:
        print(f"\n=== Ведомость остатков ===")
        print(f"Сматчено: {stock_only_matched} / xlsx={len(stocks_items)}")
        print(f"Остатки обновлены: {stock_only_changes}")
        if stock_unmatched_examples:
            print(f"Без матча в TMA: {len(stock_unmatched_examples)} (первые 10):")
            for n in stock_unmatched_examples[:10]:
                print(f"  + {n[:75]}")
    print(f"\nXLSX-товаров без матча в TMA: {len(items) - matched}")
    only_in_xlsx = [src['name'] for src in items if src['name'] not in used_xlsx]
    print(f"Примеры (первые 10):")
    for n in only_in_xlsx[:10]: print(f"  + {n[:75]}")
    print(f"\nTMA-товаров без матча в xlsx: {len(targets) - matched}")
    only_in_tma = [t['name'] for t in targets if t['id'] not in used_tma]
    for n in only_in_tma[:10]: print(f"  - {n[:75]}")


if __name__ == "__main__":
    main()
