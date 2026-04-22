"""
Админ-панель RBR Coffee Shop — расширенная версия
"""

import io
import csv
import sqlite3
import openpyxl
from datetime import datetime
from aiogram import Dispatcher, F, Bot
from aiogram.filters import Command
from aiogram.types import (
    Message, CallbackQuery, BufferedInputFile,
    InlineKeyboardMarkup, InlineKeyboardButton
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

# ─── Настройки ───────────────────────────────────────────────────────────────
ADMIN_IDS = [466755177]

# ─── FSM ─────────────────────────────────────────────────────────────────────
class AdminStates(StatesGroup):
    # Загрузка xlsx
    waiting_xlsx         = State()
    # Добавить товар
    waiting_new_name     = State()
    waiting_new_cat      = State()
    waiting_new_price    = State()
    waiting_new_stock    = State()
    waiting_new_photo    = State()
    # Редактировать товар
    waiting_edit_price   = State()
    waiting_edit_stock   = State()
    waiting_edit_photo   = State()
    # Пометки товара
    waiting_tag_product  = State()
    # Промокоды
    waiting_promo_code   = State()
    waiting_promo_disc   = State()
    waiting_promo_uses   = State()
    waiting_promo_user   = State()
    # Рассылка
    waiting_broadcast    = State()
    # Изменение цены с уведомлением
    waiting_price_notify = State()

# ─── Хелперы ─────────────────────────────────────────────────────────────────
import os
DB_PATH = os.environ.get("DB_PATH", "/app/data/shop.db")
def get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def admin_log(admin_id: int, action: str):
    con = get_db()
    con.execute(
        "INSERT INTO admin_log (admin_id, action, created_at) VALUES (?,?,?)",
        (admin_id, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    con.commit(); con.close()

def init_admin_db():
    """Создаёт новые таблицы если их нет."""
    con = get_db()
    con.executescript("""
        CREATE TABLE IF NOT EXISTS promo_codes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            code       TEXT    NOT NULL UNIQUE,
            discount   REAL    NOT NULL,
            uses_left  INTEGER NOT NULL DEFAULT 1,
            user_id    INTEGER,
            is_active  INTEGER NOT NULL DEFAULT 1,
            created_at TEXT    DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS admin_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id   INTEGER NOT NULL,
            action     TEXT    NOT NULL,
            created_at TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS wishlist (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            UNIQUE(user_id, product_id)
        );

        CREATE TABLE IF NOT EXISTS notify_when_available (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            UNIQUE(user_id, product_id)
        );
    """)

    # Миграция products — только если таблица уже существует
    tables = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]

    if "products" in tables:
        prod_cols = [r[1] for r in con.execute("PRAGMA table_info(products)").fetchall()]
        for col, typ in [
            ("tag",        "TEXT DEFAULT ''"),
            ("prev_price", "REAL DEFAULT 0"),
        ]:
            if col not in prod_cols:
                con.execute(f"ALTER TABLE products ADD COLUMN {col} {typ}")

    if "orders" in tables:
        order_cols = [r[1] for r in con.execute("PRAGMA table_info(orders)").fetchall()]
        if "promo_code" not in order_cols:
            con.execute("ALTER TABLE orders ADD COLUMN promo_code TEXT DEFAULT ''")

    con.commit(); con.close()

# ─── Главное меню ─────────────────────────────────────────────────────────────
def admin_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Загрузить xlsx",        callback_data="adm_upload")],
        [InlineKeyboardButton(text="➕ Добавить товар",        callback_data="adm_add_product")],
        [InlineKeyboardButton(text="✏️ Редактировать товар",   callback_data="adm_edit_list")],
        [InlineKeyboardButton(text="🏷 Пометки товаров",       callback_data="adm_tags_list")],
        [InlineKeyboardButton(text="❌ Удалить товар",         callback_data="adm_delete_list")],
        [InlineKeyboardButton(text="📋 Все заказы",           callback_data="adm_orders")],
        [InlineKeyboardButton(text="📊 Выгрузка заказов",     callback_data="adm_export")],
        [InlineKeyboardButton(text="👥 Клиенты",              callback_data="adm_clients")],
        [InlineKeyboardButton(text="🎁 Промокоды",            callback_data="adm_promos")],
        [InlineKeyboardButton(text="📢 Рассылка",             callback_data="adm_broadcast")],
        [InlineKeyboardButton(text="📊 Статистика",           callback_data="adm_stats")],
        [InlineKeyboardButton(text="📝 Лог действий",         callback_data="adm_log")],
    ])

def leaf_categories_keyboard(prefix: str):
    con = get_db()
    cats = con.execute("""
        SELECT c.id, c.name, p.name as parent_name
        FROM categories c LEFT JOIN categories p ON c.parent_id = p.id
        WHERE c.parent_id IS NULL
    """).fetchall()
    # Берём все категории
    all_cats = con.execute("SELECT id, name FROM categories").fetchall()
    con.close()
    buttons = [[InlineKeyboardButton(
        text=c["name"], callback_data=f"{prefix}{c['id']}"
    )] for c in all_cats]
    buttons.append([InlineKeyboardButton(text="◀ Отмена", callback_data="adm_back")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ─── /admin ───────────────────────────────────────────────────────────────────
async def admin_cmd(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.")
        return
    await message.answer("🔧 *Панель администратора*", parse_mode="Markdown",
                         reply_markup=admin_keyboard())

async def adm_back(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("🔧 *Панель администратора*", parse_mode="Markdown",
                                  reply_markup=admin_keyboard())
    await callback.answer()

# ─── Загрузка xlsx ────────────────────────────────────────────────────────────
async def adm_upload_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "📥 Отправь xlsx-файл.\n\n"
        "*Формат:* `Номенклатура | Цена | Остаток`\n"
        "Цену можно не указывать — останется старая.",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_xlsx)
    await callback.answer()

async def adm_upload_file(message: Message, state: FSMContext, bot: Bot):
    if not message.document or not message.document.file_name.endswith(".xlsx"):
        await message.answer("⚠️ Нужен файл .xlsx"); return

    file = await bot.get_file(message.document.file_id)
    buf = io.BytesIO()
    await bot.download_file(file.file_path, buf); buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        await message.answer(f"❌ Не удалось открыть: {e}")
        await state.clear(); return

    updates = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], str) or len(row[0]) < 4: continue
        name = row[0].strip()
        nums = [round(float(c)) for c in row[1:] if isinstance(c, (int, float))]
        if not nums: continue
        updates[name] = {
            "stock": int(nums[-1]),
            "price": float(nums[-2]) if len(nums) >= 2 else None
        }

    if not updates:
        await message.answer("❌ Не нашёл данных."); await state.clear(); return

    con = get_db()
    upd_stock = upd_price = 0
    skipped = []
    for name, data in updates.items():
        row = con.execute("SELECT id, price FROM products WHERE name = ?", (name,)).fetchone()
        if not row:
            row = con.execute(
                "SELECT id, price FROM products WHERE name LIKE ?", (f"%{name[:20]}%",)
            ).fetchone()
        if row:
            con.execute("UPDATE products SET stock = ? WHERE id = ?", (data["stock"], row["id"]))
            upd_stock += 1
            if data["price"] and data["price"] > 0:
                old = row["price"]
                con.execute(
                    "UPDATE products SET prev_price = ?, price = ? WHERE id = ?",
                    (old, data["price"], row["id"])
                )
                upd_price += 1
        else:
            skipped.append(name[:50])
    con.commit(); con.close()

    admin_log(message.from_user.id, f"Загружен xlsx: {upd_stock} остатков, {upd_price} цен")
    text = f"✅ Остатки: *{upd_stock}* | Цены: *{upd_price}*"
    if skipped:
        text += f"\n⚠️ Не найдено ({len(skipped)}):\n" + "\n".join(f"• {s}" for s in skipped[:10])
    await message.answer(text, parse_mode="Markdown", reply_markup=admin_keyboard())
    await state.clear()

# ─── Добавить товар ───────────────────────────────────────────────────────────
async def adm_add_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("✏️ Введи *название* нового товара:", parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_new_name); await callback.answer()

async def adm_add_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await message.answer("📂 Выбери категорию:", reply_markup=leaf_categories_keyboard("adm_newcat_"))

async def adm_add_cat(callback: CallbackQuery, state: FSMContext):
    cat_id = int(callback.data.replace("adm_newcat_", ""))
    await state.update_data(cat_id=cat_id)
    await callback.message.answer("💰 Введи *цену* (₽):", parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_new_price); await callback.answer()

async def adm_add_price(message: Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", "."))
    except ValueError:
        await message.answer("⚠️ Введи число."); return
    await state.update_data(price=price)
    await message.answer("📦 Введи *остаток*:", parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_new_stock)

async def adm_add_stock(message: Message, state: FSMContext):
    try:
        stock = int(message.text)
    except ValueError:
        await message.answer("⚠️ Введи целое число."); return
    await state.update_data(stock=stock)
    await message.answer("🖼 Ссылка на фото (URL) или *-* чтобы пропустить:", parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_new_photo)

async def adm_add_photo(message: Message, state: FSMContext):
    photo_url = None if message.text.strip() == "-" else message.text.strip()
    data = await state.get_data()
    con = get_db()
    con.execute(
        "INSERT INTO products (name, price, stock, category_id, photo_url) VALUES (?,?,?,?,?)",
        (data["name"], data["price"], data["stock"], data["cat_id"], photo_url)
    )
    con.commit(); con.close()
    admin_log(message.from_user.id, f"Добавлен товар: {data['name']}")
    await message.answer(
        f"✅ Добавлен: *{data['name']}*\n{data['price']:.0f} ₽ | {data['stock']} шт",
        parse_mode="Markdown", reply_markup=admin_keyboard()
    )
    await state.clear()

# ─── Редактировать товар ──────────────────────────────────────────────────────
async def adm_edit_list(callback: CallbackQuery):
    con = get_db()
    products = con.execute(
        "SELECT id, name, price, stock FROM products ORDER BY name LIMIT 50"
    ).fetchall()
    con.close()
    buttons = [[InlineKeyboardButton(
        text=f"{p['name'][:36]} | {p['price']:.0f}₽ | {p['stock']}",
        callback_data=f"adm_edit_{p['id']}"
    )] for p in products]
    buttons.append([InlineKeyboardButton(text="◀ Назад", callback_data="adm_back")])
    await callback.message.answer("✏️ Выбери товар:",
                                  reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

async def adm_edit_product(callback: CallbackQuery, state: FSMContext):
    product_id = int(callback.data.replace("adm_edit_", ""))
    con = get_db()
    p = con.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    con.close()
    await state.update_data(edit_product_id=product_id)
    tag_str = {"NEW": "🆕 Новинка", "EXPECTED": "⏳ Ожидается", "SALE": "📉 Снижена цена", "": "—"}.get(p["tag"] or "", "—")
    await callback.message.answer(
        f"📦 *{p['name']}*\n"
        f"Цена: {p['price']:.0f} ₽ | Остаток: {p['stock']} | Пометка: {tag_str}\n\nЧто изменить?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💰 Цену",      callback_data="adm_chg_price")],
            [InlineKeyboardButton(text="📦 Остаток",   callback_data="adm_chg_stock")],
            [InlineKeyboardButton(text="🖼 Фото",      callback_data="adm_chg_photo")],
            [InlineKeyboardButton(text="🏷 Пометку",   callback_data=f"adm_tag_{product_id}")],
            [InlineKeyboardButton(text="◀ Назад",     callback_data="adm_back")],
        ])
    )
    await callback.answer()

async def adm_chg_price_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("💰 Введи новую цену (₽):")
    await state.set_state(AdminStates.waiting_edit_price); await callback.answer()

async def adm_chg_price(message: Message, state: FSMContext, bot: Bot):
    try:
        price = float(message.text.replace(",", "."))
    except ValueError:
        await message.answer("⚠️ Введи число."); return
    data = await state.get_data()
    pid = data["edit_product_id"]
    con = get_db()
    p = con.execute("SELECT name, price FROM products WHERE id = ?", (pid,)).fetchone()
    old_price = p["price"]
    # Фиксируем старую цену и ставим пометку
    tag = "SALE" if price < old_price else ("" if price == old_price else "")
    con.execute("UPDATE products SET prev_price = ?, price = ?, tag = ? WHERE id = ?",
                (old_price, price, tag, pid))
    con.commit()

    # Уведомляем клиентов у кого в избранном
    wishers = con.execute(
        "SELECT user_id FROM wishlist WHERE product_id = ?", (pid,)
    ).fetchall()
    con.close()

    admin_log(message.from_user.id, f"Изменена цена '{p['name']}': {old_price:.0f}→{price:.0f}")
    change = "📉 снижена" if price < old_price else "📈 повышена"
    for w in wishers:
        try:
            await bot.send_message(
                w["user_id"],
                f"💰 Цена на *{p['name']}* {change}!\n{old_price:.0f} → *{price:.0f} ₽*",
                parse_mode="Markdown"
            )
        except Exception:
            pass

    await message.answer(f"✅ Цена: *{price:.0f} ₽*", parse_mode="Markdown",
                         reply_markup=admin_keyboard())
    await state.clear()

async def adm_chg_stock_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("📦 Введи новый остаток:")
    await state.set_state(AdminStates.waiting_edit_stock); await callback.answer()

async def adm_chg_stock(message: Message, state: FSMContext, bot: Bot):
    try:
        stock = int(message.text)
    except ValueError:
        await message.answer("⚠️ Введи целое число."); return
    data = await state.get_data()
    pid = data["edit_product_id"]
    con = get_db()
    old = con.execute("SELECT stock, name, tag FROM products WHERE id = ?", (pid,)).fetchone()
    con.execute("UPDATE products SET stock = ? WHERE id = ?", (stock, pid))
    # Снимаем пометку EXPECTED если товар появился
    if old["tag"] == "EXPECTED" and stock > 0:
        con.execute("UPDATE products SET tag = '' WHERE id = ?", (pid,))
    con.commit()

    # Уведомляем подписчиков на появление товара
    if old["stock"] == 0 and stock > 0:
        notifiers = con.execute(
            "SELECT user_id FROM notify_when_available WHERE product_id = ?", (pid,)
        ).fetchall()
        con.execute("DELETE FROM notify_when_available WHERE product_id = ?", (pid,))
        con.commit()
        for n in notifiers:
            try:
                await bot.send_message(
                    n["user_id"],
                    f"🔔 *{old['name']}* снова в наличии!\n\nОткрой каталог и сделай заказ ☕",
                    parse_mode="Markdown"
                )
            except Exception:
                pass
    con.close()

    admin_log(message.from_user.id, f"Остаток '{old['name']}': {old['stock']}→{stock}")
    await message.answer(f"✅ Остаток: *{stock}*", parse_mode="Markdown",
                         reply_markup=admin_keyboard())
    await state.clear()

async def adm_chg_photo_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "🖼 Отправь новую *ссылку на фото* (URL) или *-* чтобы удалить:",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_edit_photo); await callback.answer()

async def adm_chg_photo(message: Message, state: FSMContext):
    photo_url = None if message.text.strip() == "-" else message.text.strip()
    data = await state.get_data()
    con = get_db()
    con.execute("UPDATE products SET photo_url = ? WHERE id = ?",
                (photo_url, data["edit_product_id"]))
    con.commit(); con.close()
    await message.answer("✅ Фото обновлено." if photo_url else "✅ Фото удалено.",
                         reply_markup=admin_keyboard())
    await state.clear()

# ─── Пометки товаров ──────────────────────────────────────────────────────────
async def adm_tags_list(callback: CallbackQuery):
    con = get_db()
    products = con.execute(
        "SELECT id, name, tag FROM products ORDER BY name LIMIT 50"
    ).fetchall()
    con.close()
    tag_icons = {"NEW": "🆕", "EXPECTED": "⏳", "SALE": "📉", "": "·"}
    buttons = [[InlineKeyboardButton(
        text=f"{tag_icons.get(p['tag'] or '', '·')} {p['name'][:40]}",
        callback_data=f"adm_tag_{p['id']}"
    )] for p in products]
    buttons.append([InlineKeyboardButton(text="◀ Назад", callback_data="adm_back")])
    await callback.message.answer(
        "🏷 *Пометки товаров*\n\nВыбери товар:", parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()

async def adm_tag_product(callback: CallbackQuery, state: FSMContext):
    product_id = int(callback.data.replace("adm_tag_", ""))
    con = get_db()
    p = con.execute("SELECT name, tag FROM products WHERE id = ?", (product_id,)).fetchone()
    con.close()
    await state.update_data(edit_product_id=product_id)
    await callback.message.answer(
        f"🏷 *{p['name']}*\nТекущая пометка: {p['tag'] or 'нет'}\n\nВыбери новую:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🆕 Новинка",      callback_data=f"adm_settag_{product_id}_NEW")],
            [InlineKeyboardButton(text="⏳ Ожидается",    callback_data=f"adm_settag_{product_id}_EXPECTED")],
            [InlineKeyboardButton(text="📉 Снижена цена", callback_data=f"adm_settag_{product_id}_SALE")],
            [InlineKeyboardButton(text="✅ Убрать пометку", callback_data=f"adm_settag_{product_id}_NONE")],
            [InlineKeyboardButton(text="◀ Назад",        callback_data="adm_back")],
        ])
    )
    await callback.answer()

async def adm_set_tag(callback: CallbackQuery):
    parts = callback.data.replace("adm_settag_", "").split("_")
    product_id = int(parts[0])
    tag = "" if parts[1] == "NONE" else parts[1]
    con = get_db()
    p = con.execute("SELECT name FROM products WHERE id = ?", (product_id,)).fetchone()
    con.execute("UPDATE products SET tag = ? WHERE id = ?", (tag, product_id))
    con.commit(); con.close()
    admin_log(callback.from_user.id, f"Пометка '{p['name']}': {tag or 'убрана'}")
    await callback.message.answer(
        f"✅ *{p['name']}* — пометка {'«' + tag + '»' if tag else 'убрана'}",
        parse_mode="Markdown", reply_markup=admin_keyboard()
    )
    await callback.answer()

# ─── Удалить товар ────────────────────────────────────────────────────────────
async def adm_delete_list(callback: CallbackQuery):
    con = get_db()
    products = con.execute("SELECT id, name FROM products ORDER BY name LIMIT 50").fetchall()
    con.close()
    buttons = [[InlineKeyboardButton(
        text=f"❌ {p['name'][:45]}", callback_data=f"adm_del_{p['id']}"
    )] for p in products]
    buttons.append([InlineKeyboardButton(text="◀ Назад", callback_data="adm_back")])
    await callback.message.answer("❌ Выбери товар для удаления:",
                                  reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

async def adm_delete_product(callback: CallbackQuery):
    product_id = int(callback.data.replace("adm_del_", ""))
    con = get_db()
    p = con.execute("SELECT name FROM products WHERE id = ?", (product_id,)).fetchone()
    if p:
        con.execute("DELETE FROM products WHERE id = ?", (product_id,))
        con.execute("DELETE FROM cart WHERE product_id = ?", (product_id,))
        con.commit()
        admin_log(callback.from_user.id, f"Удалён товар: {p['name']}")
        await callback.message.answer(f"✅ «{p['name']}» удалён.", reply_markup=admin_keyboard())
    con.close()
    await callback.answer()

# ─── Выгрузка заказов xlsx ────────────────────────────────────────────────────
async def adm_export(callback: CallbackQuery):
    con = get_db()
    orders = con.execute("""
        SELECT o.id, o.created_at, o.name, o.phone, o.address,
               o.total, o.discount, o.status, o.promo_code,
               u.user_type, u.company_name
        FROM orders o LEFT JOIN users u ON o.user_id = u.user_id
        ORDER BY o.created_at DESC
    """).fetchall()
    items_map = {}
    all_items = con.execute("""
        SELECT oi.order_id, p.name, oi.quantity, oi.price
        FROM order_items oi JOIN products p ON oi.product_id = p.id
    """).fetchall()
    con.close()

    for it in all_items:
        items_map.setdefault(it["order_id"], []).append(it)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    headers = ["№", "Дата", "Клиент", "Телефон", "Адрес", "Сумма", "Скидка",
               "Статус", "Промокод", "Товары"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for o in orders:
        its = items_map.get(o["id"], [])
        items_str = "; ".join(f"{it['name']} {it['quantity']}шт×{it['price']:.0f}₽" for it in its)
        client = o["company_name"] or o["name"]
        ws.append([
            o["id"], o["created_at"], client, o["phone"],
            o["address"], o["total"], o["discount"] or 0,
            o["status"], o["promo_code"] or "", items_str
        ])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["J"].width = 60

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"📊 Заказы выгружены: {len(orders)} шт."
    )
    admin_log(callback.from_user.id, f"Выгружено {len(orders)} заказов в xlsx")
    await callback.answer()

# ─── Заказы ───────────────────────────────────────────────────────────────────
async def adm_orders(callback: CallbackQuery):
    con = get_db()
    orders = con.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT 20").fetchall()
    con.close()
    if not orders:
        await callback.message.answer("📋 Заказов нет.", reply_markup=admin_keyboard())
        await callback.answer(); return
    status_emoji = {"new": "🆕", "confirmed": "✅", "done": "📦", "cancelled": "❌"}
    buttons = [[InlineKeyboardButton(
        text=f"{status_emoji.get(o['status'], '❓')} №{o['id']} | {o['total']:.0f}₽ | {o['name']} | {o['created_at'][:10]}",
        callback_data=f"adm_order_{o['id']}"
    )] for o in orders]
    buttons.append([InlineKeyboardButton(text="◀ Назад", callback_data="adm_back")])
    await callback.message.answer("📋 *Последние заказы:*", parse_mode="Markdown",
                                  reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

async def adm_order_detail(callback: CallbackQuery):
    order_id = int(callback.data.replace("adm_order_", ""))
    con = get_db()
    o = con.execute("SELECT * FROM orders WHERE id = ?", (order_id,)).fetchone()
    items = con.execute("""
        SELECT oi.quantity, oi.price, p.name
        FROM order_items oi JOIN products p ON oi.product_id = p.id
        WHERE oi.order_id = ?
    """, (order_id,)).fetchall()
    con.close()
    lines = [f"📋 *Заказ №{o['id']}* [{o['status']}]\n",
             f"👤 {o['name']}  📱 {o['phone']}",
             f"🏠 {o['address']}",
             f"📅 {o['created_at']}\n"]
    for it in items:
        lines.append(f"• {it['name'][:40]}\n  {it['quantity']} × {it['price']:.0f} ₽")
    lines.append(f"\n💰 Итого: *{o['total']:.0f} ₽*")
    if o["discount"]:
        lines.append(f"🎁 Скидка: {o['discount']:.0f} ₽")
    if o["promo_code"]:
        lines.append(f"🎫 Промокод: {o['promo_code']}")
    await callback.message.answer(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"adm_status_{order_id}_confirmed")],
            [InlineKeyboardButton(text="📦 Выполнен",   callback_data=f"adm_status_{order_id}_done")],
            [InlineKeyboardButton(text="❌ Отменить",   callback_data=f"adm_status_{order_id}_cancelled")],
            [InlineKeyboardButton(text="◀ К заказам",  callback_data="adm_orders")],
        ])
    )
    await callback.answer()

async def adm_set_status(callback: CallbackQuery, bot: Bot):
    rest = callback.data[len("adm_status_"):]
    order_id_str, new_status = rest.rsplit("_", 1)
    order_id = int(order_id_str)
    con = get_db()
    con.execute("UPDATE orders SET status = ? WHERE id = ?", (new_status, order_id))
    o = con.execute("SELECT * FROM orders WHERE id = ?", (order_id,)).fetchone()
    con.commit(); con.close()
    status_text = {"confirmed": "✅ Подтверждён", "done": "📦 Выполнен",
                   "cancelled": "❌ Отменён"}.get(new_status, new_status)
    await callback.message.answer(f"Заказ №{order_id} → {status_text}",
                                  reply_markup=admin_keyboard())
    try:
        await bot.send_message(
            o["user_id"],
            f"📦 *Статус заказа №{order_id}*\n\n{status_text}", parse_mode="Markdown"
        )
    except Exception:
        pass
    admin_log(callback.from_user.id, f"Статус заказа №{order_id}: {new_status}")
    await callback.answer()

# ─── Клиенты + история ────────────────────────────────────────────────────────
async def adm_clients(callback: CallbackQuery):
    con = get_db()
    clients = con.execute("""
        SELECT u.user_id, u.name, u.phone, u.user_type, u.company_name,
               COUNT(o.id) as orders_count,
               COALESCE(SUM(o.total), 0) as total_spent,
               MAX(o.created_at) as last_order
        FROM users u
        LEFT JOIN orders o ON u.user_id = o.user_id AND o.status != 'cancelled'
        GROUP BY u.user_id
        ORDER BY total_spent DESC
        LIMIT 30
    """).fetchall()
    con.close()
    if not clients:
        await callback.message.answer("👥 Клиентов нет.", reply_markup=admin_keyboard())
        await callback.answer(); return
    buttons = [[InlineKeyboardButton(
        text=f"{'🏢' if c['user_type']=='company' else '👤'} {c['company_name'] or c['name']} | {c['orders_count']} заказов | {c['total_spent']:.0f}₽",
        callback_data=f"adm_client_{c['user_id']}"
    )] for c in clients]
    buttons.append([InlineKeyboardButton(text="◀ Назад", callback_data="adm_back")])
    await callback.message.answer(
        f"👥 *Клиенты ({len(clients)}):*", parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()

async def adm_client_detail(callback: CallbackQuery):
    user_id = int(callback.data.replace("adm_client_", ""))
    con = get_db()
    u = con.execute("SELECT * FROM users WHERE user_id = ?", (user_id,)).fetchone()
    orders = con.execute("""
        SELECT id, created_at, total, status FROM orders
        WHERE user_id = ? ORDER BY created_at DESC LIMIT 10
    """, (user_id,)).fetchall()
    con.close()

    name = u["company_name"] or u["name"]
    lines = [f"👤 *{name}*\n📱 {u['phone']}"]
    if u["user_type"] == "company":
        lines.append(f"ИНН: {u['inn']} | 📧 {u['email']}")
    lines.append(f"\n📋 *Последние заказы:*")
    status_e = {"new": "🆕", "confirmed": "✅", "done": "📦", "cancelled": "❌"}
    for o in orders:
        lines.append(f"  {status_e.get(o['status'], '?')} №{o['id']} {o['created_at'][:10]} — {o['total']:.0f} ₽")
    if not orders:
        lines.append("  Заказов нет")

    await callback.message.answer(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📢 Написать клиенту", callback_data=f"adm_msg_{user_id}")],
            [InlineKeyboardButton(text="◀ К клиентам", callback_data="adm_clients")],
        ])
    )
    await callback.answer()

# ─── Промокоды ────────────────────────────────────────────────────────────────
async def adm_promos(callback: CallbackQuery):
    con = get_db()
    promos = con.execute(
        "SELECT * FROM promo_codes ORDER BY created_at DESC LIMIT 20"
    ).fetchall()
    con.close()
    lines = ["🎁 *Промокоды:*\n"]
    for p in promos:
        status = "✅" if p["is_active"] and p["uses_left"] > 0 else "❌"
        target = f"для user {p['user_id']}" if p["user_id"] else "для всех"
        lines.append(
            f"{status} `{p['code']}` — скидка {p['discount']:.0f}% "
            f"| осталось: {p['uses_left']} | {target}"
        )
    if not promos:
        lines.append("Промокодов нет.")
    await callback.message.answer(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Создать промокод", callback_data="adm_promo_new")],
            [InlineKeyboardButton(text="◀ Назад",            callback_data="adm_back")],
        ])
    )
    await callback.answer()

async def adm_promo_new_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "🎁 Введи *код промокода* (латиницей, без пробелов):\n\nПример: `COFFEE10`",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_promo_code); await callback.answer()

async def adm_promo_code(message: Message, state: FSMContext):
    code = message.text.strip().upper()
    await state.update_data(promo_code=code)
    await message.answer("💰 Скидка в процентах (например: 10 для 10%):")
    await state.set_state(AdminStates.waiting_promo_disc)

async def adm_promo_disc(message: Message, state: FSMContext):
    try:
        disc = float(message.text.replace(",", "."))
        if not 1 <= disc <= 100: raise ValueError
    except ValueError:
        await message.answer("⚠️ Введи число от 1 до 100."); return
    await state.update_data(promo_disc=disc)
    await message.answer(
        "🔢 Сколько раз можно использовать?\n"
        "• Напиши число (например 1 для разового)\n"
        "• Или *0* для неограниченного",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_promo_uses)

async def adm_promo_uses(message: Message, state: FSMContext):
    try:
        uses = int(message.text)
        if uses < 0: raise ValueError
    except ValueError:
        await message.answer("⚠️ Введи целое число ≥ 0."); return
    await state.update_data(promo_uses=uses if uses > 0 else 999999)
    await message.answer(
        "👤 Для конкретного клиента?\n"
        "• Введи его Telegram ID\n"
        "• Или *-* чтобы промокод был для всех",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_promo_user)

async def adm_promo_user(message: Message, state: FSMContext):
    txt = message.text.strip()
    user_id = None if txt == "-" else int(txt) if txt.isdigit() else None
    data = await state.get_data()
    con = get_db()
    try:
        con.execute(
            "INSERT INTO promo_codes (code, discount, uses_left, user_id) VALUES (?,?,?,?)",
            (data["promo_code"], data["promo_disc"], data["promo_uses"], user_id)
        )
        con.commit()
        admin_log(message.from_user.id,
                  f"Создан промокод {data['promo_code']} {data['promo_disc']}% uses={data['promo_uses']}")
        target = f"для user {user_id}" if user_id else "для всех"
        await message.answer(
            f"✅ Промокод создан!\n\n"
            f"🎁 *{data['promo_code']}* — скидка {data['promo_disc']:.0f}%\n"
            f"Использований: {data['promo_uses']} | {target}",
            parse_mode="Markdown", reply_markup=admin_keyboard()
        )
    except Exception:
        await message.answer("❌ Такой код уже существует.", reply_markup=admin_keyboard())
    con.close()
    await state.clear()

# ─── Рассылка ─────────────────────────────────────────────────────────────────
async def adm_broadcast_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "📢 Введи текст рассылки.\n\n"
        "_Поддерживается Markdown: *жирный*, _курсив_, и т.д._\n"
        "Рассылка уйдёт всем зарегистрированным клиентам.",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_broadcast); await callback.answer()

async def adm_broadcast_send(message: Message, state: FSMContext, bot: Bot):
    text = message.text.strip()
    con = get_db()
    users = con.execute("SELECT user_id FROM users").fetchall()
    con.close()

    sent = failed = 0
    for u in users:
        try:
            await bot.send_message(u["user_id"], text, parse_mode="Markdown")
            sent += 1
        except Exception:
            failed += 1

    admin_log(message.from_user.id, f"Рассылка: {sent} доставлено, {failed} ошибок")
    await message.answer(
        f"📢 Рассылка завершена!\n✅ Доставлено: {sent}\n❌ Ошибок: {failed}",
        reply_markup=admin_keyboard()
    )
    await state.clear()

# ─── Лог действий ─────────────────────────────────────────────────────────────
async def adm_log_view(callback: CallbackQuery):
    con = get_db()
    logs = con.execute(
        "SELECT * FROM admin_log ORDER BY created_at DESC LIMIT 30"
    ).fetchall()
    con.close()
    if not logs:
        await callback.message.answer("📝 Лог пуст.", reply_markup=admin_keyboard())
        await callback.answer(); return
    lines = ["📝 *Лог действий:*\n"]
    for log in logs:
        lines.append(f"• {log['created_at'][:16]}: {log['action']}")
    text = "\n".join(lines)
    for i in range(0, len(text), 4000):
        await callback.message.answer(text[i:i+4000], parse_mode="Markdown")
    await callback.answer()

# ─── Статистика ───────────────────────────────────────────────────────────────
async def adm_stats(callback: CallbackQuery):
    con = get_db()
    total_orders   = con.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    total_revenue  = con.execute(
        "SELECT COALESCE(SUM(total),0) FROM orders WHERE status != 'cancelled'"
    ).fetchone()[0]
    new_orders     = con.execute("SELECT COUNT(*) FROM orders WHERE status='new'").fetchone()[0]
    total_products = con.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    total_clients  = con.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    low_stock      = con.execute("SELECT COUNT(*) FROM products WHERE stock <= 5 AND stock > 0").fetchone()[0]
    out_of_stock   = con.execute("SELECT COUNT(*) FROM products WHERE stock = 0").fetchone()[0]
    top = con.execute("""
        SELECT p.name, SUM(oi.quantity) as sold
        FROM order_items oi JOIN products p ON oi.product_id = p.id
        GROUP BY p.id ORDER BY sold DESC LIMIT 5
    """).fetchall()
    promo_used = con.execute(
        "SELECT COUNT(*) FROM orders WHERE promo_code != '' AND promo_code IS NOT NULL"
    ).fetchone()[0]
    con.close()
    lines = [
        "📊 *Статистика магазина*\n",
        f"👥 Клиентов: *{total_clients}*",
        f"📋 Заказов: *{total_orders}*  |  🆕 Новых: *{new_orders}*",
        f"💰 Выручка: *{total_revenue:.0f} ₽*",
        f"🎁 Заказов с промокодом: *{promo_used}*\n",
        f"📦 Товаров: *{total_products}*",
        f"⚠️ Заканчивается (≤5): *{low_stock}*",
        f"❌ Нет в наличии: *{out_of_stock}*",
    ]
    if top:
        lines.append("\n🏆 *Топ продаж:*")
        for i, t in enumerate(top, 1):
            lines.append(f"  {i}. {t['name'][:40]} — {t['sold']} шт")
    await callback.message.answer("\n".join(lines), parse_mode="Markdown",
                                  reply_markup=admin_keyboard())
    await callback.answer()

# ─── Уведомление о новом заказе ──────────────────────────────────────────────
async def notify_new_order(bot: Bot, order_id: int, user_name: str, total: float):
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id,
                f"🔔 *Новый заказ №{order_id}!*\n👤 {user_name}\n💰 {total:.0f} ₽\n\n/admin → Все заказы",
                parse_mode="Markdown"
            )
        except Exception:
            pass

# ─── Регистрация хендлеров ────────────────────────────────────────────────────
def register_admin_handlers(dp: Dispatcher, bot: Bot):
    init_admin_db()

    dp.message.register(admin_cmd, Command("admin"))
    dp.callback_query.register(adm_back,    F.data == "adm_back")
    dp.callback_query.register(adm_stats,   F.data == "adm_stats")
    dp.callback_query.register(adm_clients, F.data == "adm_clients")
    dp.callback_query.register(adm_client_detail, F.data.startswith("adm_client_"))
    dp.callback_query.register(adm_orders,  F.data == "adm_orders")
    dp.callback_query.register(adm_order_detail,
        F.data.startswith("adm_order_") & ~F.data.startswith("adm_orders"))
    dp.callback_query.register(lambda c: adm_set_status(c, bot), F.data.startswith("adm_status_"))
    dp.callback_query.register(adm_export,  F.data == "adm_export")

    dp.callback_query.register(adm_upload_start, F.data == "adm_upload")
    dp.message.register(lambda m, s: adm_upload_file(m, s, bot), AdminStates.waiting_xlsx)

    dp.callback_query.register(adm_add_start,   F.data == "adm_add_product")
    dp.message.register(adm_add_name,            AdminStates.waiting_new_name)
    dp.callback_query.register(adm_add_cat,      F.data.startswith("adm_newcat_"))
    dp.message.register(adm_add_price,           AdminStates.waiting_new_price)
    dp.message.register(adm_add_stock,           AdminStates.waiting_new_stock)
    dp.message.register(adm_add_photo,           AdminStates.waiting_new_photo)

    dp.callback_query.register(adm_edit_list,    F.data == "adm_edit_list")
    dp.callback_query.register(adm_edit_product, F.data.startswith("adm_edit_"))
    dp.callback_query.register(adm_chg_price_start, F.data == "adm_chg_price")
    dp.message.register(lambda m, s: adm_chg_price(m, s, bot), AdminStates.waiting_edit_price)
    dp.callback_query.register(adm_chg_stock_start, F.data == "adm_chg_stock")
    dp.message.register(lambda m, s: adm_chg_stock(m, s, bot), AdminStates.waiting_edit_stock)
    dp.callback_query.register(adm_chg_photo_start, F.data == "adm_chg_photo")
    dp.message.register(adm_chg_photo, AdminStates.waiting_edit_photo)

    dp.callback_query.register(adm_tags_list,   F.data == "adm_tags_list")
    dp.callback_query.register(adm_tag_product, F.data.startswith("adm_tag_"))
    dp.callback_query.register(adm_set_tag,     F.data.startswith("adm_settag_"))

    dp.callback_query.register(adm_delete_list,    F.data == "adm_delete_list")
    dp.callback_query.register(adm_delete_product, F.data.startswith("adm_del_"))

    dp.callback_query.register(adm_promos,         F.data == "adm_promos")
    dp.callback_query.register(adm_promo_new_start, F.data == "adm_promo_new")
    dp.message.register(adm_promo_code,  AdminStates.waiting_promo_code)
    dp.message.register(adm_promo_disc,  AdminStates.waiting_promo_disc)
    dp.message.register(adm_promo_uses,  AdminStates.waiting_promo_uses)
    dp.message.register(adm_promo_user,  AdminStates.waiting_promo_user)

    dp.callback_query.register(adm_broadcast_start, F.data == "adm_broadcast")
    dp.message.register(lambda m, s: adm_broadcast_send(m, s, bot), AdminStates.waiting_broadcast)

    dp.callback_query.register(adm_log_view, F.data == "adm_log")
