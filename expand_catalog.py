"""Расширяет TMA-каталог: добавляет товары из 'Прайс и остатки.xlsx',
которых ещё нет в products.json. Использует контекст иерархии (категория/бренд/формат)
для определения category/subcategory/fasovka.
"""
import openpyxl, json, re
from pathlib import Path

XLSX = Path("/root/projects/ai-agents-rb/BOT_TG/Прайс и остатки.xlsx")
PRODUCTS = Path("/root/projects/ai-agents-rb/BOT_TG/tma_static/products.json")


def num(v):
    if v is None or v == "": return None
    try: return float(v)
    except: return None


def normalize(s):
    s = s.upper()
    s = re.sub(r"[ЁЕ]","Е",s)
    s = re.sub(r"[^\w\s]"," ",s)
    return s


_ARTICUL_RE = re.compile(r"^(ЧА|ЧАD|АР|КФ|МК|МЛ|ТП|СР)?\d{1,5}[A-ZА-Я]?$")
TEA_NOISE = {
    "СИРОП","ТОППИНГ","СИРОПЫ","ЧАЙ","НАПИТОК","КОРДИАЛ","ПАРФЮМ",
    "ЧАИ","ЧАЯ","ЧАЙНИКОВ","ЧАЙНИКА","ЧАЙНИК","ЧАШКА","ЧАШЕК","ЧАШКУ",
    "ЧЕРНЫЙ","ЧЕРН","ЗЕЛЕНЫЙ","ЗЕЛ","ЗЕЛЕН","БЕЛЫЙ","БЕЛ","КРАСНЫЙ","ФРУКТ",
    "ФРУКТОВЫЙ","ФРУКТОВ","ТРАВЯНОЙ","ТРАВ","ЯГОДНЫЙ","ЯГОД",
    "ЛИСТОВОЙ","ЛИСТ","БАЙХОВЫЙ","БАЙХ","ПАКЕТ","ПАК","ПАКЕТИРОВАННЫЙ",
    "ПАКЕТИР","ПАКЕТИРОВ","ПИРАМИДКА","ПИРАМИДКИ","ПИРАМ","ПИР",
    "АРОМАТ","АРОМАТИЗ","АРОМАТИЗИРОВАННЫЙ","АРОМ","АРАМАТ",
    "BARLINE","BOTANIKA","ALTHAUS","АЛЬТХАУС","АЛЬТАУС",
    "NIKTEA","ROASTBERRY","RBR","HERBARISTA","SWEETSHOT","ICEDREAM",
    "GREEN","MILK","BARBACKS","CLAVIS","МОНИН","МИЛЛЕР","VEDRENNE",
    "Б","А","Л","КГ","Г","ШТ","МЛ","НА","СОЕВОЙ","ОСНОВЕ","ДЛЯ",
    "NEW","НОВ","ХИТ","ЧАЙН","ИЗ","ОТ","ПО","ВО","СО","ВКУС","ВКУСОМ",
    "БУТ","СТЕКЛ","СТ","БАНКА","БАН","КОФЕ",
}


def kw(s):
    s = normalize(s)
    out = set()
    for w in s.split():
        if len(w) < 3: continue
        if w.isdigit(): continue
        if _ARTICUL_RE.match(w): continue
        out.add(w)
    return out - TEA_NOISE


def score(a, b):
    ka, kb = kw(a), kw(b)
    if not ka or not kb: return 0.0
    return len(ka & kb) / max(len(ka), len(kb))


def slug(s):
    s = normalize(s).lower()
    s = re.sub(r"[^\w]+","_",s)
    return s[:50]


# Определение subcategory по бренду + формату (context из xlsx-иерархии)
def detect_subcat(top_section, brand, sub_format):
    """top_section: МОЛОКО/СИРОПЫ/ЧАЙ
       brand: ALTHAUS/NIKTEA/BARLINE/BOTANIKA/...
       sub_format: 'Чай для чайников 15х4' / 'Сиропы 1л' / null"""
    s = (top_section or "").upper()
    b = (brand or "").upper()
    fmt = (sub_format or "").lower()

    if "МОЛОК" in s: return "milk_main"

    if "СИРОП" in s or "ТОППИНГ" in s.lower() or "ТОППИНГ" in (sub_format or "").lower():
        if "BARLINE" in b: return "syr_barline"
        if "BOTANIKA" in b: return "syr_botanika"
        if "SWEETSHOT" in b or "СВИТШОТ" in b: return "syr_sweetshot"
        if "HERBARISTA" in b or "ГЕРБАРИСТА" in b: return "syr_herbarista"
        if "КОРДИАЛ" in b: return "syr_кордиал_и_другие"
        return "syr_other"

    if "ЧАЙ" in s:
        # ALTHAUS / NIKTEA по фасовке
        if "ALTHAUS" in b or "АЛЬТХАУС" in b:
            if "пирамид" in fmt or "pyra" in fmt: return "tea_althaus_pyr"
            if "чайник" in fmt: return "tea_althaus_pot"
            if "чашк" in fmt: return "tea_althaus_cup"
            if "лист" in fmt or "300г" in fmt or "200г" in fmt or "250 г" in fmt: return "tea_althaus_loose"
            return "tea_althaus_loose"
        if "NIKTEA" in b:
            if "пирамид" in fmt: return "tea_niktea_pyr"
            if "чайник" in fmt: return "tea_niktea_pot"
            if "чашк" in fmt: return "tea_niktea_cup"
            if "top" in fmt or "selection" in fmt: return "tea_niktea_top"
            return "tea_niktea_loose"
        if "RBR" in b: return "tea_rbr_tea_loose"
        if "REST" in b: return "tea_restoranica_loose"
        if "КИТАЙ" in b: return "tea_китайский_чай_loose"
    return None


def detect_fasovka_size(name):
    """Извлекает фасовку из имени."""
    n = name.lower()
    # Пакетики 15х4 / 20х1,75 / 25х2 / 15*2,7
    m = re.search(r"(\d+)\s*[хx*×]\s*(\d+(?:[.,]\d+)?)", n)
    if m:
        return f"{m.group(1)}х{m.group(2).replace(',','.')}г"
    # Литры/мл
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*л\b", n)
    if m: return f"{m.group(1).replace(',','.')} л"
    m = re.search(r"(\d+)\s*мл\b", n)
    if m: return f"{m.group(1)} мл"
    # Граммы
    m = re.search(r"(\d+)\s*(?:г|гр)\b", n)
    if m: return f"{m.group(1)} г"
    # Килограммы
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*кг", n)
    if m: return f"{m.group(1).replace(',','.')} кг"
    # Штуки
    m = re.search(r"(\d+)\s*шт", n)
    if m: return f"{m.group(1)} шт"
    return "1 шт"


def clean_name(raw):
    """Очищает имя из xlsx: убирает артикулы, NEW, etc."""
    s = raw
    # Убираем "ЧА016 NEW " префиксы
    s = re.sub(r"^(ЧА|ЧАD|АР|КФ|МК|ТП|СР)?\d{1,5}[A-ZА-Я]?\s*", "", s)
    s = re.sub(r"\bNEW\b\s*", "", s)
    # Убираем тяжёлые описательные хвосты после первой запятой
    parts = s.split(",")
    if len(parts) >= 2:
        # имя = первая часть; формат — последняя
        head = parts[0].strip()
        # Если есть "напиток"/"чай" в head — оставим
        return head
    return s.strip()


def parse_xlsx_with_context():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []
    section, brand, sub_fmt = "", "", ""
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 5).value
        if not name: continue
        name = str(name).strip()
        row_obj = ws.row_dimensions.get(r)
        outline = row_obj.outline_level if row_obj else 0

        if outline == 0:
            section = name; brand = ""; sub_fmt = ""; continue
        if outline == 1:
            brand = name; sub_fmt = ""; continue
        if outline == 2:
            # Это либо подформат (Чай для чайников) либо товар (если есть цена)
            price_check = num(ws.cell(r, 15).value)
            if price_check is None:
                sub_fmt = name; continue
            # есть цена → это товар (брендов где outline=2 = товар)
        # outline >= 2 + есть цена = товар
        price = num(ws.cell(r, 15).value)
        stock = num(ws.cell(r, 17).value)
        if price is None: continue
        items.append({
            "raw_name": name,
            "clean_name": clean_name(name),
            "section": section, "brand": brand, "sub_fmt": sub_fmt,
            "price": price, "stock": int(stock or 0),
        })
    return items


def main():
    items = parse_xlsx_with_context()
    print(f"Распарсено товаров (с контекстом): {len(items)}")

    data = json.loads(PRODUCTS.read_text(encoding="utf-8"))
    products = data["products"]
    targets = [p for p in products if p.get("category") != "coffee"]

    # Шаг 1: матчим существующие — обновляем цены/остатки
    pairs = []
    for src in items:
        # Используем clean_name для матча
        for t in targets:
            s = score(src["clean_name"], t["name"])
            if s > 0.0:
                pairs.append((s, src, t))
    pairs.sort(reverse=True, key=lambda x: x[0])

    used_tma = set()
    used_xlsx = set()
    matched = 0
    for s, src, t in pairs:
        if s < 0.5: break
        if t["id"] in used_tma: continue
        if id(src) in used_xlsx: continue
        used_tma.add(t["id"]); used_xlsx.add(id(src))
        if t["fasovka"]:
            t["fasovka"][0]["price"] = src["price"]
        t["stock"] = src["stock"]
        matched += 1

    # Шаг 2: добавляем недостающие
    new_subcats_needed = set()
    added = 0
    skipped_unknown = []
    for src in items:
        if id(src) in used_xlsx: continue  # уже привязали
        subcat = detect_subcat(src["section"], src["brand"], src["sub_fmt"])
        if not subcat:
            skipped_unknown.append(src["raw_name"][:60])
            continue
        # Проверим что подкатегория есть в TMA, иначе создадим
        if not any(sub["id"] == subcat for sub in data["subcategories"]):
            new_subcats_needed.add(subcat)

        category = ("milk" if subcat == "milk_main"
                    else "syrup" if subcat.startswith("syr_")
                    else "tea")
        fasovka_size = detect_fasovka_size(src["raw_name"])
        nice_name = src["clean_name"]
        if not nice_name or len(nice_name) < 4:
            nice_name = src["raw_name"][:80]

        new_id = f"{category[0]}-{slug(nice_name)}"
        # уникальность
        i = 2
        base_id = new_id
        while any(p["id"] == new_id for p in data["products"]):
            new_id = f"{base_id}_{i}"; i += 1

        item = {
            "id": new_id, "category": category, "subcategory": subcat,
            "name": nice_name, "country": src["brand"] or "",
            "roast": "", "process": "",
            "description": "",
            "recipe_e": "", "recipe_f": "",
            "fasovka": [{"size": fasovka_size, "price": src["price"]}],
            "stock": src["stock"], "tags": [],
            "photo": None,
        }
        data["products"].append(item)
        added += 1

    # Создаём недостающие подкатегории
    SUBCAT_LABELS = {
        "syr_other": "🍯 Прочие сиропы",
        "tea_niktea_top": "NIKTEA · 🌟 Top Selection",
        "tea_niktea_pyr": "NIKTEA · 🔺 Пирамидки",
        "tea_althaus_pyr": "ALTHAUS · 🔺 Пирамидки",
    }
    for sc in new_subcats_needed:
        category = ("milk" if sc == "milk_main"
                    else "syrup" if sc.startswith("syr_")
                    else "tea")
        label = SUBCAT_LABELS.get(sc, sc.replace("_", " ").title())
        data["subcategories"].append({"id": sc, "parent": category, "name": label})

    PRODUCTS.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n=== Итого ===")
    print(f"Сматчено существующих (price+stock): {matched}")
    print(f"Добавлено новых SKU: {added}")
    print(f"Создано новых подкатегорий: {len(new_subcats_needed)}")
    print(f"Не добавлено (неизвестная категория): {len(skipped_unknown)}")
    if skipped_unknown:
        print(f"Примеры:")
        for n in skipped_unknown[:6]: print(f"  ! {n}")
    print(f"\nТеперь в TMA-каталоге всего: {len(data['products'])} товаров")


if __name__ == "__main__":
    main()
