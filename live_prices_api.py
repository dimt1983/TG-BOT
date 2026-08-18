"""
live_prices_api.py — даёт TMA живые цены кофе из Bishop'овского чистового прайса.

Структура источника:
    /root/projects/ai-agents-rb/Прайсы/чистовики/Roastberry_Прайс_2026.xlsx
        Лист 'Прайс 2026': Тип | Название | Обжарка | 1кг | 10кг | 25кг | 200г | 10кг_200 | 25кг_200

Где используется:
    1. tma_static_server.py подключает endpoint /tma/api/prices
    2. TMA при загрузке делает fetch(`api/prices`) и переливает свежие цены
       поверх coffee-секции в products.json
    3. Чай / сиропы / молоко не трогаются

Кэширование: 60 секунд (mtime файла + время чтения).
"""
import json
import os
import re
import time
from pathlib import Path

# Поиск xlsx в нескольких местах: env > рядом с ботом (deploy) > синхронизированная папка
_HERE = Path(__file__).parent
_CANDIDATES = [
    Path(os.environ.get("BISHOP_PRICE_XLSX", "")) if os.environ.get("BISHOP_PRICE_XLSX") else None,
    _HERE / "tma_static" / "data" / "Roastberry_Прайс_2026.xlsx",
    Path("/root/projects/ai-agents-rb/Прайсы/чистовики/Roastberry_Прайс_2026.xlsx"),
]
PRICE_XLSX = next((p for p in _CANDIDATES if p and p.exists()), _CANDIDATES[1])

_STM_CANDIDATES = [
    Path(os.environ.get("BISHOP_STM_XLSX", "")) if os.environ.get("BISHOP_STM_XLSX") else None,
    _HERE / "tma_static" / "data" / "Roastberry_СТМ_2026.xlsx",
    Path("/root/projects/ai-agents-rb/Прайсы/чистовики/Roastberry_СТМ_2026.xlsx"),
]
STM_XLSX = next((p for p in _STM_CANDIDATES if p and p.exists()), _STM_CANDIDATES[1])

_CACHE = {"data": None, "ts": 0, "mtime": 0}
_STM_CACHE = {"data": None, "ts": 0, "mtime": 0}
_CACHE_TTL = 60  # сек


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r'[^\w\-]+', '_', s, flags=re.UNICODE)
    return s[:50]


def load_prices() -> dict:
    """Возвращает словарь:
    {
        "by_id": {"c-bittiер": {"price_1000": 1910, "price_200": 410, "discount_10kg": 1720, ...}},
        "by_name": {"БИТТЕР": {...}},
        "meta": {"course_usd": 88, "loaded_at": ts, "source": "..."}
    }
    """
    if not PRICE_XLSX.exists():
        return {"by_id": {}, "by_name": {}, "meta": {"error": "price file not found"}}

    mtime = PRICE_XLSX.stat().st_mtime
    now = time.time()

    if _CACHE["data"] and now - _CACHE["ts"] < _CACHE_TTL and _CACHE["mtime"] == mtime:
        return _CACHE["data"]

    try:
        import openpyxl
    except ImportError:
        return {"by_id": {}, "by_name": {}, "meta": {"error": "openpyxl not installed"}}

    wb = openpyxl.load_workbook(PRICE_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    by_name = {}
    course = None
    rows_processed = 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]

        # Курс
        if cells[0].lower().startswith("курс"):
            m = re.search(r'(\d+)', cells[0])
            if m:
                course = int(m.group(1))
            continue

        # Шапка
        if cells[0].lower() == "тип":
            continue

        # Это товарная строка?
        type_ = cells[0]
        name = cells[1] if len(cells) > 1 else ""
        if not name or not type_:
            continue
        # Пытаемся прочесть цены
        try:
            price_1kg = float(cells[3]) if len(cells) > 3 and cells[3] else None
            price_10kg = float(cells[4]) if len(cells) > 4 and cells[4] else None
            price_25kg = float(cells[5]) if len(cells) > 5 and cells[5] else None
            price_200 = float(cells[6]) if len(cells) > 6 and cells[6] else None
            price_200_10 = float(cells[7]) if len(cells) > 7 and cells[7] else None
            price_200_25 = float(cells[8]) if len(cells) > 8 and cells[8] else None
        except (ValueError, TypeError):
            continue

        if price_1kg is None and price_200 is None:
            continue

        roast = cells[2] if len(cells) > 2 else ""

        by_name[name.upper()] = {
            "type": type_,
            "name": name,
            "roast_tag": roast,
            "price_1000": price_1kg,
            "price_200": price_200,
            "discount_tiers": {
                "10kg_per_kg": price_10kg,
                "25kg_per_kg": price_25kg,
                "10kg_per_200g": price_200_10,
                "25kg_per_200g": price_200_25,
            },
        }
        rows_processed += 1

    data = {
        "by_id": {},
        "by_name": by_name,
        "meta": {
            "course_usd": course,
            "loaded_at": now,
            "rows": rows_processed,
            "source": str(PRICE_XLSX),
        },
    }

    _CACHE["data"] = data
    _CACHE["ts"] = now
    _CACHE["mtime"] = mtime
    return data


def _norm_words(s: str) -> set[str]:
    """Нормализованный набор слов для fuzzy-матчинга."""
    s = s.upper()
    # Убираем варианты записи / — есть строки "А/Б/В" в xlsx, разворачиваем
    s = s.replace("/", " ")
    # Убираем спецсимволы
    s = re.sub(r'[^\w\s]+', ' ', s)
    # Нормализация некоторых ключевых слов
    repl = {
        "ИРГАЧ": "ИРГАЧИФ", "ИРГАЧИФ": "ИРГАЧИФ", "ИРГАЧИФФ": "ИРГАЧИФ",
        "ГВАТЕМАЛЛА": "ГВАТЕМАЛА",
        "РУАНАДА": "РУАНДА", "РУАНДА": "РУАНДА",
        "МУТИТЕЛЛИ": "МУТЕТЕЛИ", "МУТЕТЕЛИ": "МУТЕТЕЛИ", "МУТИТЕЛИ": "МУТЕТЕЛИ",
        "БИЛОЯ": "БЕЛОЯ", "БЕЛОЯ": "БЕЛОЯ", "БЕЛОЙЯ": "БЕЛОЯ",
        "САН": "САН", "RAFAEL": "САН РАФАЕЛЬ", "РАФАЕЛЬ": "РАФАЕЛЬ",
        "BLACK": "BLACK", "ЭДИШН": "EDITION", "EDITION": "EDITION",
    }
    words = []
    for w in s.split():
        if len(w) < 3 and not w.isdigit():
            continue
        words.append(repl.get(w, w))
    # Удаляем шумовые слова
    noise = {"СУХОЙ", "МЫТЫЙ", "ХАНИ", "АНАЭРОБ", "ТОП", "ГР", "1", "2", "3", "4",
             "ЭДИШН", "EDITION", "ОБЖАРКА", "ТЕМНАЯ", "СВЕТЛАЯ", "ДАРК", "DARK",
             "ЭСПРЕССО", "ЭСПР", "ФИЛЬТР", "КГ",
             "НАТУР", "НАТУРАЛ", "НАТУРАЛЬНАЯ", "НАТУРАЛЬНЫЙ"}
    return set(words) - noise


def _grade(name):
    """Номер грейда из имени: 'гр.4' / 'Gr 2' / 'гр2' → '4'/'2'. Нужен чтобы
    не путать сорта разного грейда (Сидамо Гр.4 ≠ Сидамо гр.2)."""
    m = re.search(r'(?:гр|gr)\.?\s*(\d)', name, re.IGNORECASE)
    return m.group(1) if m else None


def _match_score(tma_name: str, price_name: str) -> float:
    # a — карточка магазина (длинное имя: сорт + обработка/фасовка/линейка),
    # b — позиция прайса (короткое каноничное имя).
    a = _norm_words(tma_name)
    b = _norm_words(price_name)
    if not a or not b:
        return 0.0
    # Разный грейд одного сорта — разный товар, не матчим.
    ga, gb = _grade(tma_name), _grade(price_name)
    if ga and gb and ga != gb:
        return 0.0
    inter = len(a & b)
    sym = inter / max(len(a), len(b))
    # Directional containment: если ВСЕ значимые слова прайсовой позиции есть в
    # карточке — это матч, даже когда карточка длиннее (несёт «мытая», «FILTER»,
    # «нат.анаэроб», «NH-7» — их нет в прайсе, и они рушили симметричный score).
    # Гейт inter>=2 отсекает ложные срабатывания по одному слову-стране
    # (иначе любая «Кения …» сматчила бы «Кения АА», где значимое слово одно).
    contain = inter / len(b) if inter >= 2 else 0.0
    return max(sym, contain)


# Категории кофе, к которым применяется живой прайс из xlsx. Миграция catalog_v2
# (май 2026) разнесла бывшую единую 'coffee' на espresso/filter/black/borshch —
# старый фильтр `== "coffee"` после этого пропускал ВСЕ товары, живые цены не
# применялись (у всех карточек _price_source=None). 'coffee' — для совместимости.
COFFEE_CATEGORIES = frozenset({"coffee", "espresso", "filter", "black", "borshch"})


def _find_price_row(tma_name: str, by_name: dict):
    """Позиция прайса под карточку магазина: точное имя → первое слово → fuzzy.

    Вынесено отдельно, чтобы основной прайс и СТМ подбирались одинаково:
    разойдись они — клиент получил бы цену от чужой позиции."""
    name = (tma_name or "").upper().strip()
    if not name:
        return None
    match = by_name.get(name)
    if match:
        return match
    match = by_name.get(name.split()[0])
    if match:
        return match
    best_key, best_score = None, 0.0
    for k in by_name.keys():
        sc = _match_score(name, k)
        if sc > best_score:
            best_score, best_key = sc, k
    return by_name[best_key] if best_score >= 0.75 else None


def load_stm_prices() -> dict:
    """Прайс СТМ: {by_name: {ИМЯ: {price_1000, price_200}}, meta}.

    Лист: Тип | Название | Обжарка | Базовый (1 кг) | СТМ (1 кг) |
          Базовый (200 г) | СТМ (200 г)
    Берём именно СТМ-колонки — базовые нужны только глазами сверить."""
    if not STM_XLSX.exists():
        return {"by_name": {}, "meta": {"error": "stm price file not found"}}

    mtime = STM_XLSX.stat().st_mtime
    now = time.time()
    if _STM_CACHE["data"] and now - _STM_CACHE["ts"] < _CACHE_TTL and _STM_CACHE["mtime"] == mtime:
        return _STM_CACHE["data"]

    try:
        import openpyxl
    except ImportError:
        return {"by_name": {}, "meta": {"error": "openpyxl not installed"}}

    wb = openpyxl.load_workbook(STM_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_name, rows = {}, 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) < 5 or not cells[0] or not cells[1]:
            continue
        if cells[0].lower() == "тип":
            continue
        try:
            stm_1kg = float(cells[4]) if len(cells) > 4 and cells[4] else None
            stm_200 = float(cells[6]) if len(cells) > 6 and cells[6] else None
        except (ValueError, TypeError):
            continue
        if stm_1kg is None and stm_200 is None:
            continue
        by_name[cells[1].upper()] = {
            "name": cells[1],
            "price_1000": stm_1kg,
            "price_200": stm_200,
        }
        rows += 1

    data = {"by_name": by_name,
            "meta": {"loaded_at": now, "rows": rows, "source": str(STM_XLSX)}}
    _STM_CACHE.update(data=data, ts=now, mtime=mtime)
    return data


def stm_prices_for_products(tma_products: list[dict]) -> tuple[dict, dict]:
    """{id карточки: {размер фасовки: цена СТМ}} + статистика.

    Размеры берём как они записаны в карточке, чтобы витрине и серверу не
    пришлось нормализовать их по-разному."""
    prices = load_stm_prices()
    by_name = prices["by_name"]
    out, matched, unmatched = {}, 0, []
    if not by_name:
        return out, {"matched": 0, "unmatched_count": 0, "meta": prices["meta"]}

    for p in tma_products:
        if p.get("category") not in COFFEE_CATEGORIES:
            continue
        match = _find_price_row(p.get("name") or "", by_name)
        if not match:
            unmatched.append(p.get("name"))
            continue
        sizes = {}
        for fa in (p.get("fasovka") or []):
            size = fa.get("size") or ""
            if "1 кг" in size and match.get("price_1000"):
                sizes[size] = int(match["price_1000"] + 0.5)
            elif "200 г" in size and match.get("price_200"):
                sizes[size] = int(match["price_200"] + 0.5)
        if sizes:
            out[p["id"]] = sizes
            matched += 1
    return out, {"matched": matched, "unmatched_count": len(unmatched),
                 "unmatched": unmatched[:20], "meta": prices["meta"]}


def merge_into_products(tma_products: list[dict]) -> tuple[list[dict], dict]:
    """Применяет цены к coffee-товарам TMA. Возвращает (обновлённые товары, статистика)."""
    prices = load_prices()
    by_name = prices["by_name"]
    matched = 0
    unmatched = []
    out = []

    for p in tma_products:
        if p.get("category") not in COFFEE_CATEGORIES:
            out.append(p)
            continue

        # Карточка помечена как ручная цена — live-merge её не трогает.
        # Это позволяет Bishop'у/Дмитрию задавать цены, которых нет в xlsx
        # (например пересчитанные по новому коэффициенту), и они не перетрутся
        # старыми значениями fuzzy-матча.
        if p.get("_price_locked"):
            out.append(p)
            continue

        # Порог fuzzy высокий (0.75) — иначе позиции, которых нет в xlsx,
        # сваливаются на чужие имена: «Колумбия Андино» vs «Колумбия Андино
        # мытый» это 1.0, а «Эфиопия Иргачиф гр.4» vs «Иргач Адада гр.1» — 0.67.
        match = _find_price_row(p["name"], by_name)

        if match:
            # Перепишем цены в fasovka
            new_fasovka = []
            for fa in p.get("fasovka", []):
                if "1 кг" in fa["size"] and match.get("price_1000"):
                    fa = dict(fa)
                    fa["price"] = int(match["price_1000"] + 0.5)  # округл. вверх: 2847.5→2848, как на витрине
                elif "200 г" in fa["size"] and match.get("price_200"):
                    fa = dict(fa)
                    fa["price"] = int(match["price_200"] + 0.5)
                new_fasovka.append(fa)
            p = dict(p)
            p["fasovka"] = new_fasovka
            p["_price_source"] = "bishop"
            p["_discount_tiers"] = match.get("discount_tiers")
            matched += 1
        else:
            unmatched.append(p["name"])

        out.append(p)

    return out, {
        "matched": matched,
        "unmatched_count": len(unmatched),
        "unmatched": unmatched[:20],
        "meta": prices["meta"],
    }


# === Тест из CLI ===
if __name__ == "__main__":
    import sys
    products_path = Path(__file__).parent / "tma_static" / "products.json"
    if not products_path.exists():
        print(f"Не найден {products_path}")
        sys.exit(1)
    tma = json.loads(products_path.read_text(encoding="utf-8"))
    updated, stats = merge_into_products(tma["products"])
    print(f"Матч: {stats['matched']} / {sum(1 for p in tma['products'] if p.get('category') == 'coffee')}")
    print(f"Не сматчено: {stats['unmatched_count']}")
    if stats['unmatched']:
        print("Примеры:")
        for n in stats['unmatched'][:10]:
            print(f"  - {n}")
    print(f"\nМета: {stats['meta']}")

    # Покажем пример
    sample = next((p for p in updated if p.get("_price_source") == "bishop"), None)
    if sample:
        print(f"\nПример обновлённого:")
        print(f"  {sample['name']}")
        for fa in sample.get('fasovka', []):
            print(f"  → {fa['size']}: {fa['price']} ₽")
        if sample.get('_discount_tiers'):
            print(f"  Опт-цены: {sample['_discount_tiers']}")
